"""CPOR U4 — export builder + API contract tests (no cip)."""

from __future__ import annotations

import hashlib
from datetime import date
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

from openpyxl import load_workbook

from app.services.cpor.export_xlsx import RESELLER_HEADERS, build_cpor_case_workbook_bytes
from app.services.cpor.pivot import build_case_pivot, is_voided_line
from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)


def test_reseller_headers_frozen():
    assert RESELLER_HEADERS == (
        "Case Code",
        "Case Name",
        "Customer",
        "Promotion Type",
        "Window Start",
        "Window End",
        "Status",
        "Version",
        "ROE",
        "SKU",
        "Product Name",
        "Product Line",
        "Distributor",
        "POD Quarter",
        "SOH",
        "SRP",
        "VAT Rate",
        "Dealer Margin %",
        "Dealer Price",
        "Cost Basis",
        "Cost Source",
        "Support/Unit",
        "Estimate Qty",
        "Cap Qty",
        "Ttl Support",
        "Support USD",
        "Ttl Support USD",
        "Result Qty",
        "Ttl Result",
        "Ttl Result USD",
        "Remark",
        "Flags",
    )


def test_is_voided_line():
    assert is_voided_line(SimpleNamespace(remark="ok [voided]", estimate_qty=0))
    assert not is_voided_line(SimpleNamespace(remark=None, estimate_qty=20))


def test_builder_renders_stored_workbook_numbers():
    case = SimpleNamespace(
        id=1,
        case_code="C26C99999",
        case_name="Evetech Jan Sell Out Support",
        customer_id=7,
        promotion_type="Sell out PP",
        window_start=date(2026, 1, 1),
        window_end=date(2026, 1, 31),
        status="proposed",
        export_version=1,
        roe_snapshot=18.5,
    )
    line = SimpleNamespace(
        id=1,
        case_id=1,
        product_id=9,
        distributor_id=None,
        pod_quarter="26Q1",
        soh_snapshot=10,
        srp=13999,
        vat_rate=0.15,
        dealer_margin_pct=0.15,
        dealer_price=10347.09,
        cost_basis=9585.07,
        cost_source="manual",
        cost_evidence_json={"flags": []},
        support_unit=250.96,
        estimate_qty=20,
        cap_qty=None,
        ttl_support=5019.2,
        support_usd=13.5654,
        ttl_support_usd=271.308,
        result_qty=49,
        ttl_result=12297.18,
        ttl_result_usd=664.7124,
        remark=None,
    )
    voided = SimpleNamespace(
        id=2,
        case_id=1,
        product_id=9,
        distributor_id=None,
        pod_quarter="26Q1",
        soh_snapshot=0,
        srp=1,
        vat_rate=0.15,
        dealer_margin_pct=0.15,
        dealer_price=1,
        cost_basis=1,
        cost_source="manual",
        cost_evidence_json={},
        support_unit=0,
        estimate_qty=0,
        cap_qty=None,
        ttl_support=0,
        support_usd=0,
        ttl_support_usd=0,
        result_qty=None,
        ttl_result=None,
        ttl_result_usd=None,
        remark="x [voided]",
    )
    cust = SimpleNamespace(id=7, code="EV", name="Evetech")
    prod = SimpleNamespace(id=9, sku="SKU-1", name="Notebook", product_line="NB")

    session = MagicMock()
    session.get = MagicMock(side_effect=lambda model, pk: case if pk == 1 else cust)
    session.scalars = MagicMock(
        side_effect=[
            MagicMock(all=MagicMock(return_value=[line, voided])),  # lines
            MagicMock(all=MagicMock(return_value=[prod])),  # products
        ]
    )

    data, digest, meta = build_cpor_case_workbook_bytes(session, 1)
    assert meta["line_count"] == 1
    assert "no_distributor" in meta["flags_present"]
    assert hashlib.sha256(data).hexdigest() == digest

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == ["Reseller", "USD Pivot"]
    headers = [c.value for c in wb["Reseller"][1]]
    assert tuple(headers) == RESELLER_HEADERS
    row = [c.value for c in wb["Reseller"][2]]
    # Dealer Price / Support/Unit / Ttl Result columns
    assert row[headers.index("Case Name")] == "Evetech Jan Sell Out Support"
    assert row[headers.index("Product Line")] == "NB"
    assert row[headers.index("Dealer Price")] == 10347.09
    assert row[headers.index("Support/Unit")] == 250.96
    assert row[headers.index("Ttl Result")] == 12297.18
    assert row[headers.index("Ttl Support USD")] == 271.308
    assert row[headers.index("Ttl Result USD")] == 664.7124
    assert wb["Reseller"].max_row == 2  # header + 1 line (voided excluded)


def test_pivot_helper_totals():
    case = SimpleNamespace(id=1, case_code="X", roe_snapshot=18.5)
    line = SimpleNamespace(
        product_id=1,
        pod_quarter="26Q1",
        support_usd=10.0,
        ttl_support_usd=50.0,
        estimate_qty=5,
        remark=None,
    )
    products = {1: SimpleNamespace(product_line="NB")}
    session = MagicMock()
    pivot = build_case_pivot(session, case, [line], products)
    assert pivot["grand_total_usd"] == 50.0
    assert pivot["cells"]["26Q1"]["NB"] == 50.0


def test_export_routes_registered():
    paths = {getattr(r, "path", None) for r in app.routes}
    assert "/api/v1/cpor/cases/{case_id}/export" in paths
    assert "/api/v1/cpor/cases/{case_id}/exports" in paths
    assert "/api/v1/cpor/cases/{case_id}/exports/{version}/file" in paths
