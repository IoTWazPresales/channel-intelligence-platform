"""Focused tests for generic CPOR payment-evidence parser (no DB)."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.cpor.payment_evidence.parser import parse_payment_workbook
from app.services.cpor.payment_evidence.profile_defaults import asus_pending_report_profile_dict


KEN = Path(r"c:\Users\warren_eliason\Downloads\NewCPOR_PendingReport_101508 (59).xlsx")


@pytest.mark.skipif(not KEN.exists(), reason="Ken Pending Report not on this machine")
def test_parse_asus_pending_report_profile_against_ken_file():
    data = KEN.read_bytes()
    result = parse_payment_workbook(data, profile=asus_pending_report_profile_dict())
    assert result["ok"] is True, result.get("blocking_errors")
    assert result["row_count"] >= 1000
    row = result["rows"][0]
    assert row["external_case_code"]
    assert "amount" in row
    assert row["currency_code"]
    # source_key prefers case + credit_note when present
    assert str(row["source_key"]).startswith("pay:")


def test_parse_requires_case_code_column():
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws.append(["Subject", "Amount"])
    ws.append(["no case", 10])
    buf = BytesIO()
    wb.save(buf)
    result = parse_payment_workbook(buf.getvalue(), profile=asus_pending_report_profile_dict())
    assert result["ok"] is False
    assert any(e.get("code") == "missing_case_code_column" for e in result["blocking_errors"])


def test_parse_latest_comment_not_collapsed_into_subject():
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws.append(["Case ID", "Subject", "Latest Comment", "Payment Status", "Deduction No", "CN Status"])
    ws.append(["C80", "eShop Y25 Oct Sell Out Support", "overclaim 75 units", "To_Be_Clarified", "1", "Open"])
    buf = BytesIO()
    wb.save(buf)
    result = parse_payment_workbook(buf.getvalue(), profile=asus_pending_report_profile_dict())
    assert result["ok"] is True
    row = result["rows"][0]
    assert row["description"] == "eShop Y25 Oct Sell Out Support"
    assert row["flags_json"]["latest_comment"] == "overclaim 75 units"
    assert row["flags_json"]["deduction_no"] == "1"
    assert row["flags_json"]["cn_status_raw"] == "Open"


def test_parse_optional_owed_amount_file_into_flags():
    wb = Workbook()
    ws = wb.active
    ws.title = "Detail"
    ws.append(["Case ID", "Amount", "Payment Status", "Owed"])
    ws.append(["C99", 10, "Closed", 250])
    buf = BytesIO()
    wb.save(buf)
    result = parse_payment_workbook(buf.getvalue(), profile=asus_pending_report_profile_dict())
    assert result["ok"] is True
    assert result["rows"][0]["flags_json"]["owed_amount_file"] == 250.0
    assert result["rows"][0]["amount"] == 10.0
