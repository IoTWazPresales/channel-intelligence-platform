"""DSI multi-sheet workbook helpers."""

from __future__ import annotations

import io

import pandas as pd
import pytest

from app.services.imports.dsi_workbook import (
    DSI_SINGLE_SHEET_KEY,
    build_combined_dsi_dataframe,
    build_dsi_workbook_structure,
    is_nested_dsi_field_mapping,
    load_dsi_workbook_sheet_frames,
    _sheet_looks_dsi_mappable,
)
from app.services.imports.dsi_batch import normalized_header_signature


def test_single_sheet_mapping_unchanged_shape() -> None:
    assert is_nested_dsi_field_mapping({"Qty": "quantity_sold", "SKU": "product_identifier"}) is False


def test_nested_mapping_detected() -> None:
    assert is_nested_dsi_field_mapping({"Sheet1": {"Qty": "quantity_sold"}}) is True


def test_two_sheet_workbook_combined_stages_both_canonicals() -> None:
    sell = pd.DataFrame(
        {
            "Dist": ["D1"],
            "SKU": ["P1"],
            "Qty": [5],
            "TxDate": ["2024-01-01"],
            "Cust": ["C1"],
        }
    )
    soh = pd.DataFrame(
        {
            "Dist": ["D1"],
            "SKU": ["P1"],
            "SOH": [100],
            "Snap": ["2024-01-31"],
        }
    )
    frames = [
        (
            "Sellout",
            sell,
            {
                "Dist": "distributor_token",
                "SKU": "product_identifier",
                "Qty": "quantity_sold",
                "TxDate": "transaction_date",
                "Cust": "customer_dealer_token",
            },
        ),
        (
            "SOH",
            soh,
            {
                "Dist": "distributor_token",
                "SKU": "product_identifier",
                "SOH": "stock_on_hand",
                "Snap": "snapshot_date",
            },
        ),
    ]
    combined, mapping, skipped = build_combined_dsi_dataframe(frames)
    assert skipped == []
    assert len(combined) == 2
    assert "quantity_sold" in combined.columns
    assert "stock_on_hand" in combined.columns
    assert mapping["quantity_sold"] == "quantity_sold"
    assert mapping["stock_on_hand"] == "stock_on_hand"


def test_extra_unmapped_sheet_skipped() -> None:
    frames = [
        (
            "Data",
            pd.DataFrame({"Dist": ["D1"], "SKU": ["P1"], "Qty": [1]}),
            {"Dist": "distributor_token", "SKU": "product_identifier", "Qty": "quantity_sold"},
        ),
        ("Notes", pd.DataFrame({"A": ["ignore"]}), {}),
    ]
    _combined, _mapping, skipped = build_combined_dsi_dataframe(frames)
    assert any(s["sheet_name"] == "Notes" for s in skipped)


def test_combine_allows_missing_distributor_when_file_stamps() -> None:
    """Confirmed per-file stamps — sheets need product_id, not Dist column."""
    frames = [
        (
            "Sellout",
            pd.DataFrame({"SKU": ["P1"], "Qty": [2]}),
            {"SKU": "product_identifier", "Qty": "quantity_sold"},
            "mustek.xlsx",
        ),
    ]
    empty, _mapping, skipped_strict = build_combined_dsi_dataframe(frames)
    assert empty.empty
    assert any(s["reason"] == "not_mapped_or_missing_distributor" for s in skipped_strict)

    combined, mapping, skipped = build_combined_dsi_dataframe(
        frames,
        require_distributor_column=False,
    )
    assert skipped == []
    assert len(combined) == 1
    assert "distributor_token" not in combined.columns
    assert mapping["product_identifier"] == "product_identifier"
    assert combined["_dsi_source_file"].iloc[0] == "mustek.xlsx"


def test_xlsx_loads_multiple_sheets() -> None:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        pd.DataFrame({"A": [1]}).to_excel(writer, sheet_name="Sellout", index=False)
        pd.DataFrame({"B": [2]}).to_excel(writer, sheet_name="Inventory", index=False)
    frames = load_dsi_workbook_sheet_frames("test.xlsx", bio.getvalue())
    assert len(frames) == 2
    assert all(len(t) == 3 for t in frames)
    structure = build_dsi_workbook_structure("test.xlsx", bio.getvalue())
    assert structure["multi_sheet"] is True
    assert structure["sheet_count"] == 2


def _banner_xlsx_bytes(*, extra_blank_rows: int = 0) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sellout"
    row = 1
    for _ in range(extra_blank_rows):
        ws.cell(row=row, column=1, value="")
        row += 1
    ws.cell(row=row, column=1, value="PINNACLE SELLOUT REPORT")
    row += 1
    ws.cell(row=row, column=1, value="Confidential")
    row += 1
    headers = ["Distributor", "SKU", "Qty", "Date", "Customer"]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col, value=h)
    row += 1
    for col, v in enumerate(["DIST-01", "SKU-ALPHA-01", 2, "2024-06-01", "Dealer Zed"], start=1):
        ws.cell(row=row, column=col, value=v)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_banner_xlsx_sniffs_real_header_row() -> None:
    raw = _banner_xlsx_bytes()
    frames = load_dsi_workbook_sheet_frames("banner.xlsx", raw)
    assert len(frames) == 1
    _sn, df, header_row = frames[0]
    assert header_row > 0
    assert _sheet_looks_dsi_mappable(df)
    cols = {str(c).strip().lower() for c in df.columns}
    assert "distributor" in cols
    assert "sku" in cols
    structure = build_dsi_workbook_structure("banner.xlsx", raw)
    assert structure["sheets"][0]["header_row"] == header_row
    sig, _, _, unm, reason = normalized_header_signature("banner.xlsx", raw)
    assert not unm
    assert reason is None
    assert sig != "unmappable"


def test_banner_csv_sniffs_real_header_row() -> None:
    csv = (
        "MUSTEK SELLOUT WK24\n"
        "Generated report\n"
        "Distributor,SKU,Qty,Date,Customer\n"
        "DIST-01,SKU-ALPHA-01,1,2024-06-08,Dealer Zed\n"
    ).encode("utf-8")
    frames = load_dsi_workbook_sheet_frames("mustek.csv", csv)
    _sn, df, header_row = frames[0]
    assert header_row > 0
    assert _sheet_looks_dsi_mappable(df)
    sig, _, _, unm, _ = normalized_header_signature("mustek.csv", csv)
    assert not unm and sig != "unmappable"


def test_asus_weekly_sellout_form_header_at_row_19() -> None:
    """Real PINNACLE/MUSTEK sellouts: form block then table header near row 19."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sell Out"
    ws["A5"] = "Subject: ASUS Weekly Sell Out Report"
    ws["A7"] = "Method"
    ws["B7"] = "SYS"
    ws["A8"] = "Company Name"
    ws["B8"] = "MUSTEK-ZA-C"
    ws["A9"] = "Application Date"
    ws["B9"] = "2026W24"
    headers = [
        "Invoice Date",
        "Invoice No.",
        "Customer Code (Dealer Code)",
        "Dealer Name",
        "Model Name",
        "ASUS Part No.",
        "Quantity",
        "Unit Price",
    ]
    for col, h in enumerate(headers, start=1):
        ws.cell(row=20, column=col, value=h)  # openpyxl 1-based → pandas header index 19
    for col, v in enumerate(
        ["2026-06-08", "1385729", "864966", "Hbits", "TP3407", "90NB14Y1", 1, 15304],
        start=1,
    ):
        ws.cell(row=21, column=col, value=v)
    bio = io.BytesIO()
    wb.save(bio)
    raw = bio.getvalue()
    frames = load_dsi_workbook_sheet_frames("MUSTEK_SELLOUT WK24.xlsx", raw)
    _sn, df, header_row = frames[0]
    assert header_row == 19
    assert _sheet_looks_dsi_mappable(df)
    assert "Invoice Date" in list(df.columns)
    sig, _, _, unm, reason = normalized_header_signature("MUSTEK_SELLOUT WK24.xlsx", raw)
    assert not unm and reason is None and sig != "unmappable"


def test_asus_weekly_sellout_csv_header_at_row_19() -> None:
    rows = [""] * 19  # rows 0..18 form/blank
    rows[4] = "Subject: ASUS Weekly Sell Out Report"
    rows.append(
        "Invoice Date,Invoice No.,Customer Code (Dealer Code),Dealer Name,"
        "Model Name,ASUS Part No.,Quantity,Unit Price"
    )  # row 19
    rows.append("2026-07-08,INV1,C1,Dealer,MODEL,90NB,1,100")
    csv = ("\n".join(rows) + "\n").encode("utf-8")
    frames = load_dsi_workbook_sheet_frames("Sellout-PINNACLE.csv", csv)
    _sn, df, header_row = frames[0]
    assert header_row == 19
    assert _sheet_looks_dsi_mappable(df)
    sig, _, _, unm, _ = normalized_header_signature("Sellout-PINNACLE.csv", csv)
    assert not unm and sig != "unmappable"


def test_clean_xlsx_header_row_zero_unchanged() -> None:
    clean = pd.DataFrame(
        {
            "Distributor": ["DIST-01"],
            "SKU": ["SKU-ALPHA-01"],
            "Qty": [1],
            "Date": ["2024-06-01"],
            "Customer": ["C1"],
        }
    )
    bio = io.BytesIO()
    clean.to_excel(bio, index=False, engine="openpyxl")
    raw = bio.getvalue()
    frames = load_dsi_workbook_sheet_frames("clean.xlsx", raw)
    _sn, df, header_row = frames[0]
    assert header_row == 0
    assert list(df.columns) == list(clean.columns)


def test_banner_offset_stability_same_signature() -> None:
    a = _banner_xlsx_bytes(extra_blank_rows=0)
    b = _banner_xlsx_bytes(extra_blank_rows=2)
    sig_a, _, _, unm_a, _ = normalized_header_signature("a.xlsx", a)
    sig_b, _, _, unm_b, _ = normalized_header_signature("b.xlsx", b)
    assert not unm_a and not unm_b
    assert sig_a == sig_b


def test_non_dsi_finance_sheet_stays_unmappable_with_reason() -> None:
    bio = io.BytesIO()
    pd.DataFrame({"Account": ["1000"], "Balance": [12.5], "Memo": ["x"]}).to_excel(
        bio, index=False, engine="openpyxl"
    )
    sig, _, _, unm, reason = normalized_header_signature("finance.xlsx", bio.getvalue())
    assert unm
    assert sig == "unmappable"
    assert reason == "no_dsi_headers"


def test_multisheet_cover_plus_banner_dsi() -> None:
    from openpyxl import Workbook

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "Notes only"
    cover["B1"] = "Ignore"
    sell = wb.create_sheet("Sellout")
    sell["A1"] = "Title banner"
    sell["A2"] = "Distributor"
    sell["B2"] = "SKU"
    sell["C2"] = "Qty"
    sell["D2"] = "Date"
    sell["A3"] = "DIST-01"
    sell["B3"] = "SKU-ALPHA-01"
    sell["C3"] = 1
    sell["D3"] = "2024-06-01"
    bio = io.BytesIO()
    wb.save(bio)
    structure = build_dsi_workbook_structure("mixed.xlsx", bio.getvalue())
    mappable = [s for s in structure["sheets"] if s.get("dsi_mappable")]
    assert len(mappable) >= 1
    assert any(s.get("header_row", 0) > 0 for s in mappable)
