"""CPOR H1 — historical workbook parse + validate (no cip writes)."""

from __future__ import annotations

import io
from datetime import date, datetime, timedelta

from openpyxl import Workbook

from app.services.cpor.historical_import import (
    asus_default_profile_dict,
    parse_and_validate_historical_workbook,
    parse_historical_workbook,
)
from app.services.cpor.historical_import.validate import parity_check_row
from app.services.imports.template_definitions import IMPORT_TEMPLATE_ROWS


def _excel_serial(d: date) -> int:
    return (datetime(d.year, d.month, d.day) - datetime(1899, 12, 30)).days


def _build_fixture_workbook() -> bytes:
    """Minimal ASUS-shaped workbook: banner row + header + two data rows per sheet."""
    wb = Workbook()
    # Disti
    ws = wb.active
    ws.title = "Disti Sell out"
    ws.append([None] * 10 + [1, 2, 3])  # banner
    ws.append(
        [
            "Remark",
            "Case ID",
            "Case Name",
            "Promotion Type",
            "Start From",
            "End on",
            "Status",
            "POD Quarter",
            "Product Line",
            "Distributor",
            "Dealer/Retailer",
            "Sales Model Name",
            "Model2",
            "Model3",
            "SOH",
            "Estimate Qty",
            "Result",
            "SRP ",
            " VAT ",
            " Dealer/Store Margin",
            " Dealer Price ",
            "Rebate",
            "Net Price",
            "Disti Margin",
            " Disti Price ",
            "Disti System Cost",
            "DAP Invoiced",
            "Disti ROE",
            "New DAP",
            "New Disti/Customer Cost",
            " Support  P/Unit ZAR",
            " Ttl Support ",
            "ROE",
            " Support P/Unit USD",
            "Ttl Support USD",
            "Ttl Support Result USD",
            "Ttl Support Result ZAR",
        ]
    )
    start = _excel_serial(date(2023, 1, 1))
    end = _excel_serial(date(2023, 1, 31))
    # Row matching known waterfall fixture: SRP 13999, VAT 0.15, margin 0.15 → dealer 10347.09
    # cost 9585.07 → support 250.96; result 49 → ttl_result ~12297.18
    ws.append(
        [
            None,
            "C23C15833",
            "Computer Mania Jan Sell Out Support",
            "Sell out PP",
            start,
            end,
            "Approved",
            "2023Q3",
            "NB",
            "Mustek",
            "Computer mania",
            "E1504FA-O516512B0W",
            None,
            None,
            59,
            20,
            49,
            13999,
            0.15,
            0.15,
            10347.08695652174,
            0.03,
            10036.67,
            0.07,
            9334.10714347826,
            9585.07,
            None,
            0,
            504.82,
            9334.10714347826,
            250.962856521739,
            5019.25713043478,
            18.49,
            13.572896512803625,
            271.45793025607253,
            665.0719291273776,
            12297.179969565212,
        ]
    )

    # Reseller
    ws2 = wb.create_sheet("Reseller Sell out")
    ws2.append([None] * 5)
    ws2.append(
        [
            "Remark",
            "Case ID",
            "Case Name",
            "Promotion Type",
            "Start From",
            "End on",
            "Case Status",
            "Quarter",
            "Product Line",
            "Distributor",
            "Dealer/Retailer",
            "Sales Model Name",
            "Model Name",
            "Model Name",
            "SOH",
            "Estimate Qty",
            "Result",
            "SRP ",
            " VAT ",
            "Dealer Margin",
            " Dealer Price ",
            "Rebate",
            "Net Price",
            "Margin",
            "Dealer Net Price ",
            "Dealer System Cost",
            "DAP Invoiced",
            "Disti ROE",
            "New DAP",
            "New Disti/Customer Cost",
            " Support  P/Unit ZAR",
            " Ttl Support ",
            "ROE",
            " Support P/Unit USD",
            "Ttl Support USD",
            "Ttl Support Result ($)",
            "Ttl Support Result Rand",
        ]
    )
    # Sell-Through with result → settled
    ws2.append(
        [
            None,
            "C23C16002",
            "Computer Mania Jan Sell-thru Support",
            "Sell-Through PP",
            start,
            end,
            "Approved",
            "2023Q3",
            "NB",
            "Mustek",
            "Computer Mania",
            "M1502QA-58512S0W",
            None,
            None,
            403,
            10,
            40,
            8999,
            0.15,
            0.15,
            6651.434782608696,
            None,
            6651.434782608696,
            None,
            6651.434782608696,
            7760.13,
            None,
            0,
            350.44,
            6651.434782608696,
            1108.6952173913041,
            11086.952173913041,
            18.98,
            58.413868144958066,
            584.1386814495806,
            2336.5547257983226,
            44347.808695652166,
        ]
    )
    # Under discussion → skipped
    ws2.append(
        [
            None,
            "C23C99999",
            "Draft discussion",
            "Sell-Through PP",
            start,
            end,
            "Under Discussion",
            "2023Q3",
            "NB",
            "Mustek",
            "Takealot",
            "UX8402VU-OI71610B0W",
            None,
            None,
            1,
            1,
            0,
            1000,
            0.15,
            0.15,
            739.13,
            None,
            739.13,
            None,
            739.13,
            800,
            None,
            0,
            None,
            None,
            60.87,
            60.87,
            18,
            3.38,
            3.38,
            0,
            0,
        ]
    )
    wb.create_sheet("Sell out Pivot")  # ignored

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_template_slug_registered():
    slugs = {r["slug"] for r in IMPORT_TEMPLATE_ROWS}
    assert "cpor_historical_cases" in slugs


def test_asus_default_profile_has_sheet_roles():
    p = asus_default_profile_dict()
    assert p["sheet_roles_json"]["Disti Sell out"] == "disti"
    assert p["sheet_roles_json"]["Reseller Sell out"] == "reseller"
    assert p["header_row_index"] == 1


def test_parse_fixture_workbook():
    data = _build_fixture_workbook()
    parsed = parse_historical_workbook(data)
    assert not parsed["blocking_errors"]
    assert parsed["row_count"] == 3
    codes = {r["case_code"] for r in parsed["rows"]}
    assert "C23C15833" in codes
    assert "C23C16002" in codes
    disti = next(r for r in parsed["rows"] if r["case_code"] == "C23C15833")
    assert disti["channel"] == "disti"
    assert disti["lifecycle_status"] == "settled"
    assert disti["result_qty"] == 49
    reseller = next(r for r in parsed["rows"] if r["case_code"] == "C23C16002")
    assert reseller["channel"] == "reseller"
    assert reseller["lifecycle_status"] == "settled"
    skipped = next(r for r in parsed["rows"] if r["case_code"] == "C23C99999")
    assert skipped["skip_apply"] is True
    assert "under_discussion_skipped" in skipped["flags_json"]["flags"]


def test_parse_and_validate_fixture():
    data = _build_fixture_workbook()
    result = parse_and_validate_historical_workbook(data)
    assert result["ok"] is True
    assert result["validation"]["apply_candidate_count"] == 2
    assert result["validation"]["skipped_count"] == 1
    assert "C23C15833" in result["validation"]["cases"]


def test_parity_check_reseller_fixture_row():
    # Cost above dealer price so support_unit > 0 (reseller channel math)
    row = {
        "channel": "reseller",
        "srp": 13999,
        "vat_rate": 0.15,
        "dealer_margin_pct": 0.15,
        "cost_basis": 11000,
        "estimate_qty": 20,
        "result_qty": 49,
        "roe_snapshot": 18.49,
        "support_unit": 652.91304347826,  # 11000 - 10347.08695652174
        "ttl_result": 31992.73913043474,  # support * 49
    }
    flags = parity_check_row(row)
    assert "waterfall_parity_variance" not in flags


def test_parity_check_disti_uses_funded_price():
    row = {
        "channel": "disti",
        "srp": 13999,
        "vat_rate": 0.15,
        "dealer_margin_pct": 0.15,
        "cost_basis": 9585.07,
        "dealer_price": 9334.10714347826,  # Disti Price
        "estimate_qty": 20,
        "result_qty": 49,
        "support_unit": 250.96285652174,
    }
    flags = parity_check_row(row)
    assert "waterfall_parity_variance" not in flags
