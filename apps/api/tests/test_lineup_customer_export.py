"""Unit tests for the per-customer lineup slice XLSX export."""

import asyncio
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock

import pytest
from openpyxl import load_workbook

from app.services.commercial_planner.lineup_customer_export import (
    LineupExportNotFoundError,
    build_customer_lineup_slice_xlsx,
)

_CHAIN = {
    "inputs": {
        "srp_inc_vat_local": 1150.0,
        "vat_rate_pct": 0.15,
        "dealer_margin_pct": 0.20,
        "rebate_pct": 0.03,
        "distributor_margin_pct": 0.08,
        "import_tax_pct": 0.0,
        "roe_local_per_cost_currency": 18.0,
    },
    "outputs": {
        "calc_srp_ex_vat_local": 1000.0,
        "calc_dealer_price_local": 800.0,
        "calc_net_price_local": 776.0,
        "calc_disti_cost_local": 713.92,
        "calc_dap_cost_currency": 39.6622,
        "calc_profit_per_unit": None,
        "calc_profit_total": None,
    },
}


def _line():
    return SimpleNamespace(
        quantity_units=10,
        msrp_local=1150,
        sku_raw="SKU-ALPHA-01",
        model_raw="ModelX",
        calc_dap_cost_currency=39.6622,
        calc_profit_total=None,
        pricing_chain_json=_CHAIN,
        customer_feedback="Need better price",
    )


def _db_with_rows(rows):
    db = MagicMock()
    case = SimpleNamespace(
        period_label="26Q1", product_line="NB", iteration_number=2, currency_code="ZAR"
    )
    customer = SimpleNamespace(id=1001, code="CUST-1001", name="Acme")
    db.get = AsyncMock(side_effect=[case, customer])
    exec_result = MagicMock()
    exec_result.all.return_value = rows
    db.execute = AsyncMock(return_value=exec_result)
    return db


def test_export_builds_xlsx_with_pricing_chain():
    rows = [(_line(), "SKU-ALPHA-01", "Alpha NB", "ModelX", "Disti A")]
    db = _db_with_rows(rows)
    data, filename, n = asyncio.run(build_customer_lineup_slice_xlsx(db, 7, 1001))
    assert n == 1
    assert data[:2] == b"PK"
    assert filename == "lineup_case7_CUST-1001_26Q1.xlsx"
    ws = load_workbook(BytesIO(data)).active
    cells = [c for row in ws.iter_rows(values_only=True) for c in (row or [])]
    assert "Calculated DAP (cost ccy)" in cells
    assert 39.6622 in cells  # calc DAP value
    assert 800.0 in cells  # dealer price from chain outputs
    assert "Need better price" in cells  # customer feedback


def test_export_empty_slice_still_writes_headers():
    db = _db_with_rows([])
    data, _filename, n = asyncio.run(build_customer_lineup_slice_xlsx(db, 7, 1001))
    assert n == 0
    ws = load_workbook(BytesIO(data)).active
    cells = [c for row in ws.iter_rows(values_only=True) for c in (row or [])]
    assert "SKU" in cells


def test_export_missing_case_raises():
    db = MagicMock()
    db.get = AsyncMock(return_value=None)
    with pytest.raises(LineupExportNotFoundError):
        asyncio.run(build_customer_lineup_slice_xlsx(db, 999, 1001))
