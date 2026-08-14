from datetime import date
from io import BytesIO

from openpyxl import load_workbook

from app.services.lineup.export_apply import (
    half_year_period_starts,
    net_requirement_to_csv,
    net_requirement_to_xlsx_bytes,
)


def test_half_year_1h():
    slots = half_year_period_starts(2026, 1)
    assert slots == [(date(2026, 1, 1), "26Q1"), (date(2026, 4, 1), "26Q2")]


def test_half_year_2h():
    slots = half_year_period_starts(2026, 2)
    assert slots == [(date(2026, 7, 1), "26Q3"), (date(2026, 10, 1), "26Q4")]


def test_csv_header():
    csv = net_requirement_to_csv({"rows": [{"distributor_id": 1, "product_id": 2, "net_requirement": 5}]})
    assert "net_requirement" in csv.splitlines()[0]
    assert "1,2" in csv.replace(" ", "")


def test_xlsx_workbook_sheets(tmp_path, monkeypatch):
    from app.services import commercial_tenant_profile as profile

    monkeypatch.setattr(profile, "_tenant_profiles_dir", lambda: tmp_path)
    payload = {
        "rows": [
            {
                "distributor_id": 1,
                "product_id": 2,
                "business_unit": "NB",
                "forecast_demand": 10,
                "bias_adjusted_forecast": 11,
                "channel_stock": 1,
                "in_transit": 0,
                "target_cover_units": 4,
                "net_requirement": 13,
                "weekly_velocity": 1,
                "bias_factor": 1.1,
            }
        ]
    }
    draft = [
        {
            "Customer Code": "ACME",
            "Customer Name": "Acme",
            "SKU": "SKU-1",
            "Product Name": "P",
            "Period Label": "26Q2",
            "Period Start": "2026-04-01",
            "Planned Qty": 5,
            "Distributor ID": 1,
            "Product ID": 2,
            "Business Unit": "NB",
            "Forecast Demand": 10,
            "Bias Adjusted Forecast": 11,
            "Channel Stock": 1,
            "In Transit": 0,
            "Target Cover": 4,
            "Net Requirement": 13,
            "Notes": "test",
        }
    ]
    raw = net_requirement_to_xlsx_bytes(payload, draft_rows=draft)
    wb = load_workbook(BytesIO(raw))
    assert wb.sheetnames == ["NetRequirement", "DraftLineup"]
    assert wb["NetRequirement"]["A1"].value == "distributor_id"
    assert wb["DraftLineup"]["A1"].value == "Customer Code"
    assert wb["DraftLineup"]["A2"].value == "ACME"


def test_xlsx_draft_sheet_uses_tenant_column_map(tmp_path, monkeypatch):
    from app.services import commercial_tenant_profile as profile

    monkeypatch.setattr(profile, "_tenant_profiles_dir", lambda: tmp_path)
    profile.save_tenant_profile_overrides(
        "default",
        {
            "lineup_export_columns": [
                {"field": "sku", "header": "Item"},
                {"field": "planned_qty", "header": "Qty"},
                {"field": "customer_code", "header": "Cust"},
            ]
        },
    )
    payload = {"rows": []}
    draft = [{"sku": "SKU-1", "planned_qty": 9, "customer_code": "ACME"}]
    raw = net_requirement_to_xlsx_bytes(payload, draft_rows=draft)
    wb = load_workbook(BytesIO(raw))
    assert wb["DraftLineup"]["A1"].value == "Item"
    assert wb["DraftLineup"]["B1"].value == "Qty"
    assert wb["DraftLineup"]["C1"].value == "Cust"
    assert wb["DraftLineup"]["A2"].value == "SKU-1"
    assert wb["DraftLineup"]["B2"].value == 9
    assert wb["DraftLineup"]["C2"].value == "ACME"
