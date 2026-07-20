"""Parse CPOR historical tracking workbooks into staging-shaped rows (H1)."""

from __future__ import annotations

import hashlib
import io
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

from app.services.cpor.historical_import.profile_defaults import (
    asus_default_profile_dict,
    normalize_header,
)

# openpyxl stylesheet workaround for some ASUS workbooks with duplicate gradient stops
import openpyxl.reader.excel as _excel_mod

_excel_mod.apply_stylesheet = lambda *a, **k: None  # type: ignore[assignment]


def _load_workbook(data: bytes):
    from openpyxl.reader.excel import load_workbook

    return load_workbook(io.BytesIO(data), read_only=True, data_only=True, keep_links=False)


def _excel_serial_to_date(val: Any) -> date | None:
    if val is None or val == "" or val == "#VALUE!":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, (int, float)):
        try:
            n = float(val)
        except (TypeError, ValueError):
            return None
        if n > 30000:  # Excel serial
            return (datetime(1899, 12, 30) + timedelta(days=n)).date()
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nan", "none", "#value!"):
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _num(val: Any) -> Decimal | None:
    if val is None or val == "" or val == "#VALUE!":
        return None
    if isinstance(val, Decimal):
        return val
    try:
        if isinstance(val, float) and val != val:  # NaN
            return None
        return Decimal(str(val).strip().replace(",", "").replace("\xa0", ""))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _str(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).replace("\xa0", " ").strip()
    if not s or s.lower() in ("nan", "none", "#value!"):
        return None
    return s


def historical_source_key(
    *,
    channel: str,
    case_code: str,
    sales_model: str | None,
    distributor: str | None,
    pod_quarter: str | None,
    sheet_name: str,
    row_number: int,
) -> str:
    raw = "|".join(
        [
            channel,
            case_code.strip().upper(),
            (sales_model or "").strip().upper(),
            (distributor or "").strip().upper(),
            (pod_quarter or "").strip().upper(),
            sheet_name.strip().lower(),
            str(row_number),
        ]
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:64]


def _resolve_column_index(headers: list[str], aliases: list[str]) -> int | None:
    norm_headers = [normalize_header(h) for h in headers]
    for alias in aliases:
        target = normalize_header(alias)
        for i, h in enumerate(norm_headers):
            if h == target:
                return i
    return None


def _build_index(headers: list[str], column_map: dict[str, list[str]]) -> dict[str, int]:
    out: dict[str, int] = {}
    for canon, aliases in column_map.items():
        idx = _resolve_column_index(headers, aliases)
        if idx is not None:
            out[canon] = idx
    return out


def _cell(vals: list[Any], idx_map: dict[str, int], key: str) -> Any:
    i = idx_map.get(key)
    if i is None or i >= len(vals):
        return None
    return vals[i]


def parse_historical_workbook(
    data: bytes,
    *,
    profile: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Parse workbook bytes → staging-shaped rows (no DB).

    Returns {rows, sheet_summaries, profile_code, blocking_errors}.
    """
    prof = profile or asus_default_profile_dict()
    sheet_roles: dict[str, str] = dict(prof.get("sheet_roles_json") or {})
    column_map: dict[str, list[str]] = dict(prof.get("column_map_json") or {})
    header_row_index = int(prof.get("header_row_index", 1))
    value_maps: dict[str, Any] = dict(prof.get("value_maps_json") or {})
    status_map: dict[str, str] = {
        str(k).lower(): str(v) for k, v in (value_maps.get("status") or {}).items()
    }
    promo_map: dict[str, str] = {
        str(k).lower(): str(v) for k, v in (value_maps.get("promotion_type") or {}).items()
    }
    under_discussion_policy = str(value_maps.get("under_discussion_policy") or "skip_with_flag")

    wb = _load_workbook(data)
    rows_out: list[dict[str, Any]] = []
    sheet_summaries: list[dict[str, Any]] = []
    blocking: list[dict[str, str]] = []

    try:
        for sheet_name in wb.sheetnames:
            role = sheet_roles.get(sheet_name)
            if role is None:
                # Unknown sheet — ignore unless explicitly mapped
                sheet_summaries.append(
                    {"sheet": sheet_name, "role": "unmapped_ignored", "rows": 0}
                )
                continue
            if role == "ignore":
                sheet_summaries.append({"sheet": sheet_name, "role": "ignore", "rows": 0})
                continue
            if role not in ("disti", "reseller"):
                blocking.append(
                    {
                        "code": "bad_sheet_role",
                        "message": f"Sheet {sheet_name!r} role {role!r} invalid",
                    }
                )
                continue

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if len(all_rows) <= header_row_index:
                sheet_summaries.append(
                    {"sheet": sheet_name, "role": role, "rows": 0, "flag": "no_header"}
                )
                continue
            headers = [
                str(h).replace("\xa0", " ").strip() if h is not None else f"col_{i}"
                for i, h in enumerate(all_rows[header_row_index])
            ]
            idx_map = _build_index(headers, column_map)
            if "case_code" not in idx_map:
                blocking.append(
                    {
                        "code": "missing_case_code_column",
                        "message": f"Sheet {sheet_name!r}: Case ID column not mapped",
                    }
                )
                continue

            n = 0
            for rnum, raw in enumerate(all_rows[header_row_index + 1 :], start=header_row_index + 2):
                vals = list(raw)
                if not any(v is not None and str(v).strip() not in ("", "None") for v in vals):
                    continue
                case_code = _str(_cell(vals, idx_map, "case_code"))
                if not case_code:
                    continue
                n += 1
                raw_dict = {
                    headers[i]: vals[i] if i < len(vals) else None for i in range(len(headers))
                }
                status_raw = _str(_cell(vals, idx_map, "status"))
                status_norm = status_map.get((status_raw or "").lower(), "unknown")
                promo_raw = _str(_cell(vals, idx_map, "promotion_type"))
                promo_norm = promo_map.get((promo_raw or "").lower(), promo_raw)

                result_qty = _num(_cell(vals, idx_map, "result_qty"))
                estimate_qty = _num(_cell(vals, idx_map, "estimate_qty"))
                flags: list[str] = []
                skip_apply = False
                lifecycle: str | None = None

                if status_norm == "cancelled":
                    lifecycle = "cancelled"
                elif status_norm == "under_discussion":
                    if under_discussion_policy == "skip_with_flag":
                        skip_apply = True
                        flags.append("under_discussion_skipped")
                    lifecycle = "draft"
                elif status_norm == "approved":
                    if result_qty is not None and result_qty != 0:
                        lifecycle = "settled"
                    else:
                        lifecycle = "ended"
                        flags.append("no_result")
                else:
                    flags.append("unknown_status")
                    skip_apply = True

                # Prefer dealer_price; disti sheet funds off Disti Price when present
                dealer_price = _num(_cell(vals, idx_map, "dealer_price"))
                disti_price = _num(_cell(vals, idx_map, "disti_price"))
                cost_basis = _num(_cell(vals, idx_map, "cost_basis"))
                if role == "disti" and disti_price is not None:
                    funded_price = disti_price
                else:
                    funded_price = dealer_price if dealer_price is not None else disti_price

                snapshot_extra = {
                    k: _str(_cell(vals, idx_map, k)) or _num(_cell(vals, idx_map, k))
                    for k in (
                        "rebate_pct",
                        "disti_margin_pct",
                        "disti_price",
                        "dap_invoiced",
                        "disti_roe",
                        "new_dap",
                        "new_disti_customer_cost",
                    )
                    if k in idx_map
                }

                sales_model = _str(_cell(vals, idx_map, "sales_model_token"))
                distributor = _str(_cell(vals, idx_map, "distributor_token"))
                pod_quarter = _str(_cell(vals, idx_map, "pod_quarter"))
                sk = historical_source_key(
                    channel=role,
                    case_code=case_code,
                    sales_model=sales_model,
                    distributor=distributor,
                    pod_quarter=pod_quarter,
                    sheet_name=sheet_name,
                    row_number=rnum,
                )

                rows_out.append(
                    {
                        "source_key": sk,
                        "source_row_number": rnum,
                        "sheet_name": sheet_name,
                        "channel": role,
                        "case_code": case_code,
                        "case_name": _str(_cell(vals, idx_map, "case_name")),
                        "promotion_type_raw": promo_raw,
                        "promotion_type": promo_norm,
                        "window_start": _excel_serial_to_date(_cell(vals, idx_map, "window_start")),
                        "window_end": _excel_serial_to_date(_cell(vals, idx_map, "window_end")),
                        "status_raw": status_raw,
                        "lifecycle_status": lifecycle,
                        "pod_quarter": pod_quarter,
                        "product_line": _str(_cell(vals, idx_map, "product_line")),
                        "distributor_token": distributor,
                        "customer_token": _str(_cell(vals, idx_map, "customer_token")),
                        "sales_model_token": sales_model,
                        "soh_snapshot": float(x) if (x := _num(_cell(vals, idx_map, "soh_snapshot"))) is not None else None,
                        "estimate_qty": float(estimate_qty) if estimate_qty is not None else None,
                        "result_qty": float(result_qty) if result_qty is not None else None,
                        "srp": float(x) if (x := _num(_cell(vals, idx_map, "srp"))) is not None else None,
                        "vat_rate": float(x) if (x := _num(_cell(vals, idx_map, "vat_rate"))) is not None else None,
                        "dealer_margin_pct": float(x)
                        if (x := _num(_cell(vals, idx_map, "dealer_margin_pct"))) is not None
                        else None,
                        "cost_basis": float(cost_basis) if cost_basis is not None else None,
                        "dealer_price": float(funded_price) if funded_price is not None else None,
                        "support_unit": float(x)
                        if (x := _num(_cell(vals, idx_map, "support_unit"))) is not None
                        else None,
                        "ttl_support": float(x)
                        if (x := _num(_cell(vals, idx_map, "ttl_support"))) is not None
                        else None,
                        "roe_snapshot": float(x)
                        if (x := _num(_cell(vals, idx_map, "roe_snapshot"))) is not None
                        else None,
                        "support_usd": float(x)
                        if (x := _num(_cell(vals, idx_map, "support_usd"))) is not None
                        else None,
                        "ttl_support_usd": float(x)
                        if (x := _num(_cell(vals, idx_map, "ttl_support_usd"))) is not None
                        else None,
                        "ttl_result": float(x)
                        if (x := _num(_cell(vals, idx_map, "ttl_result"))) is not None
                        else None,
                        "ttl_result_usd": float(x)
                        if (x := _num(_cell(vals, idx_map, "ttl_result_usd"))) is not None
                        else None,
                        "remark": _str(_cell(vals, idx_map, "remark")),
                        "skip_apply": skip_apply,
                        "flags_json": {"flags": flags},
                        "source_snapshot_json": snapshot_extra,
                        "raw_source_row": {
                            k: (
                                v.isoformat()
                                if isinstance(v, (date, datetime))
                                else (float(v) if isinstance(v, Decimal) else v)
                            )
                            for k, v in raw_dict.items()
                        },
                    }
                )
            sheet_summaries.append({"sheet": sheet_name, "role": role, "rows": n})
    finally:
        wb.close()

    return {
        "rows": rows_out,
        "sheet_summaries": sheet_summaries,
        "profile_code": prof.get("profile_code"),
        "blocking_errors": blocking,
        "row_count": len(rows_out),
    }
