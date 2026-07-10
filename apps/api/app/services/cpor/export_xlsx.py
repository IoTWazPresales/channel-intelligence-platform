"""CPOR case XLSX builder — Reseller layout + USD pivot (spec §5).

Renders stored computed columns only — no waterfall arithmetic.
"""

from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.cpor import CporCase, CporCaseLine
from app.models.dimensions import DimCustomer, DimDistributor, DimProduct
from app.services.cpor.pivot import build_case_pivot, is_voided_line

# Presentation contract (Warren-confirmed 2026-07-09) — Reseller sheet header row
RESELLER_HEADERS: tuple[str, ...] = (
    "Case Code",
    "Case Name",
    "Customer",
    "Promotion Type",
    "Window Start",
    "Window End",
    "Status",
    "Version",
    "ROE",
    "SKU",
    "Product Name",
    "Product Line",
    "Distributor",
    "POD Quarter",
    "SOH",
    "SRP",
    "VAT Rate",
    "Dealer Margin %",
    "Dealer Price",
    "Cost Basis",
    "Cost Source",
    "Support/Unit",
    "Estimate Qty",
    "Cap Qty",
    "Ttl Support",
    "Support USD",
    "Ttl Support USD",
    "Result Qty",
    "Ttl Result",
    "Ttl Result USD",
    "Remark",
    "Flags",
)


def _f(v: Any) -> float | str | None:
    if v is None:
        return None
    if isinstance(v, str):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return str(v)


def _line_flags(line: CporCaseLine) -> list[str]:
    flags: list[str] = []
    if line.distributor_id is None:
        flags.append("no_distributor")
    if line.cost_basis is None:
        flags.append("no_cost_basis")
    ev = line.cost_evidence_json or {}
    for f in ev.get("flags") or []:
        if f not in flags:
            flags.append(str(f))
    return flags


def build_cpor_case_workbook_bytes(session: Session, case_id: int) -> tuple[bytes, str, dict[str, Any]]:
    """Return (xlsx_bytes, sha256_hex, meta). Raises ValueError if no exportable lines."""
    case = session.get(CporCase, case_id)
    if case is None:
        raise ValueError("case_not_found")
    cust = session.get(DimCustomer, case.customer_id)
    lines = list(session.scalars(select(CporCaseLine).where(CporCaseLine.case_id == case.id)).all())
    exportable = [l for l in lines if not is_voided_line(l) and float(l.estimate_qty or 0) != 0.0]
    if not exportable:
        raise ValueError("no_exportable_lines")

    product_ids = list({int(l.product_id) for l in lines})
    products = {
        int(p.id): p
        for p in session.scalars(select(DimProduct).where(DimProduct.id.in_(product_ids))).all()
    }
    dist_ids = [int(l.distributor_id) for l in exportable if l.distributor_id is not None]
    dists = {
        int(d.id): d
        for d in session.scalars(select(DimDistributor).where(DimDistributor.id.in_(dist_ids))).all()
    } if dist_ids else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Reseller"
    ws.append(list(RESELLER_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    all_flags: list[str] = []
    if case.roe_snapshot is None:
        all_flags.append("missing_roe")

    customer_label = f"{cust.code} — {cust.name}" if cust else str(case.customer_id)
    for line in exportable:
        prod = products.get(int(line.product_id))
        dist = dists.get(int(line.distributor_id)) if line.distributor_id else None
        flags = _line_flags(line)
        all_flags.extend(flags)
        dist_label = ""
        if dist:
            dist_label = f"{dist.code} — {dist.name}"
        elif line.distributor_id is None:
            dist_label = ""  # blank + no_distributor flag
        ws.append(
            [
                case.case_code,
                case.case_name,
                customer_label,
                case.promotion_type,
                case.window_start.isoformat() if case.window_start else None,
                case.window_end.isoformat() if case.window_end else None,
                case.status,
                case.export_version,
                _f(case.roe_snapshot),
                prod.sku if prod else None,
                prod.name if prod else None,
                prod.product_line if prod else None,
                dist_label,
                line.pod_quarter,
                _f(line.soh_snapshot),
                _f(line.srp),
                _f(line.vat_rate),
                _f(line.dealer_margin_pct),
                _f(line.dealer_price),
                _f(line.cost_basis),
                line.cost_source,
                _f(line.support_unit),
                _f(line.estimate_qty),
                _f(line.cap_qty),
                _f(line.ttl_support),
                _f(line.support_usd),
                _f(line.ttl_support_usd),
                _f(line.result_qty),
                _f(line.ttl_result),
                _f(line.ttl_result_usd),
                line.remark,
                ", ".join(flags),
            ]
        )

    pivot = build_case_pivot(session, case, lines, products)
    ws2 = wb.create_sheet("USD Pivot")
    product_lines = sorted(pivot["col_totals"].keys())
    ws2.append(["POD Quarter", *product_lines, "Row Total"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.freeze_panes = "B2"
    for pq in sorted(pivot["cells"].keys()):
        row = [pq]
        for pl in product_lines:
            row.append(pivot["cells"].get(pq, {}).get(pl, 0.0))
        row.append(pivot["row_totals"].get(pq, 0.0))
        ws2.append(row)
    total_row = ["Column Total"]
    for pl in product_lines:
        total_row.append(pivot["col_totals"].get(pl, 0.0))
    total_row.append(pivot["grand_total_usd"])
    ws2.append(total_row)
    if pivot["missing_roe"]:
        ws2.append([])
        ws2.append(["NOTE: missing_roe — USD values may be blank/zero on lines without ROE."])

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    digest = hashlib.sha256(data).hexdigest()
    meta = {
        "line_count": len(exportable),
        "flags_present": sorted(set(all_flags)),
        "export_version": int(case.export_version or 1),
        "case_code": case.case_code,
        "pivot_grand_total_usd": pivot["grand_total_usd"],
        "missing_roe": pivot["missing_roe"],
    }
    return data, digest, meta
