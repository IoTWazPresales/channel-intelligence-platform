"""Parse tenant payment / CN workbooks into canonical row dicts."""

from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from openpyxl import load_workbook

from app.services.cpor.payment_evidence.profile_defaults import (
    asus_pending_report_profile_dict,
    normalize_header,
)


def _cell_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    s = str(value).strip()
    return s or None


def _as_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00")).date()
    except ValueError:
        return None


def _as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def _build_index(headers: list[str], column_map: dict[str, list[str]]) -> dict[str, int]:
    norm_headers = [normalize_header(h) for h in headers]
    out: dict[str, int] = {}
    for canon, aliases in column_map.items():
        for alias in aliases:
            want = normalize_header(alias)
            for i, h in enumerate(norm_headers):
                if h == want:
                    out[canon] = i
                    break
            if canon in out:
                break
    return out


def _normalize_payment_status(raw: str | None, value_map: dict[str, str]) -> str | None:
    if not raw:
        return None
    key = normalize_header(raw)
    if key in value_map:
        return value_map[key]
    # strip leading "N. " ordinal prefixes common in some finance exports
    cleaned = key
    if ". " in cleaned:
        cleaned = cleaned.split(". ", 1)[-1].strip()
    return value_map.get(cleaned) or key.replace(" ", "_")


def _source_key(*, case_code: str, credit_note_id: str | None, row_number: int, sheet: str) -> str:
    cn = (credit_note_id or "").strip()
    if cn:
        return f"pay:{case_code}:{cn}"
    return f"pay:{case_code}:row:{sheet}:{row_number}"


def parse_payment_workbook(data: bytes, profile: dict[str, Any] | None = None) -> dict[str, Any]:
    """Return {ok, rows, sheet_summaries, profile_code, blocking_errors}."""
    prof = profile or asus_pending_report_profile_dict()
    sheet_roles: dict[str, str] = dict(prof.get("sheet_roles_json") or {})
    column_map: dict[str, list[str]] = dict(prof.get("column_map_json") or {})
    value_maps: dict[str, Any] = dict(prof.get("value_maps_json") or {})
    status_map: dict[str, str] = dict(value_maps.get("payment_status") or {})
    header_row_index = int(prof.get("header_row_index") or 1)

    from io import BytesIO

    # Scoped stylesheet workaround (same class of openpyxl duplicate-stop issues as historical)
    import openpyxl.styles.stylesheet as _excel_mod

    original_apply = _excel_mod.apply_stylesheet
    _excel_mod.apply_stylesheet = lambda *a, **k: None  # type: ignore[assignment]
    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    finally:
        _excel_mod.apply_stylesheet = original_apply

    rows_out: list[dict[str, Any]] = []
    sheet_summaries: list[dict[str, Any]] = []
    blocking_errors: list[dict[str, str]] = []

    role_sheets = [n for n, role in sheet_roles.items() if role == "payment_evidence"]
    if not role_sheets:
        # Fallback: first sheet if profile forgot roles (generic single-sheet feeds)
        role_sheets = list(wb.sheetnames[:1])

    for sheet_name in wb.sheetnames:
        role = sheet_roles.get(sheet_name)
        if role is None:
            sheet_summaries.append({"sheet": sheet_name, "role": "unmapped_ignored", "rows": 0})
            continue
        if role == "ignore":
            sheet_summaries.append({"sheet": sheet_name, "role": "ignore", "rows": 0})
            continue
        if role != "payment_evidence":
            blocking_errors.append(
                {"code": "bad_sheet_role", "message": f"Sheet {sheet_name!r} role {role!r} invalid"}
            )
            continue

        ws = wb[sheet_name]
        all_rows = ws.iter_rows(values_only=True)
        headers: list[str] | None = None
        for i, raw in enumerate(all_rows, start=1):
            if i < header_row_index:
                continue
            if i == header_row_index:
                headers = [str(c) if c is not None else "" for c in raw]
                continue
            if headers is None:
                continue
            if raw is None or all(c is None or str(c).strip() == "" for c in raw):
                continue

            idx = _build_index(headers, column_map)
            if "external_case_code" not in idx:
                blocking_errors.append(
                    {
                        "code": "missing_case_code_column",
                        "message": f"Sheet {sheet_name!r}: external_case_code column not mapped",
                    }
                )
                break

            def get(canon: str) -> Any:
                j = idx.get(canon)
                if j is None or j >= len(raw):
                    return None
                return raw[j]

            case_code = _cell_str(get("external_case_code"))
            if not case_code:
                continue

            credit_note_id = _cell_str(get("credit_note_id"))
            payment_status_raw = _cell_str(get("payment_status_raw"))
            amount = _as_decimal(get("amount"))
            currency = _cell_str(get("currency_code")) or "ZAR"
            customer_token = _cell_str(get("customer_token"))
            distributor_token = _cell_str(get("distributor_token"))
            description = _cell_str(get("description"))
            case_status_raw = _cell_str(get("case_status_raw"))
            payment_date = _as_date(get("payment_date"))
            window_start = _as_date(get("window_start"))
            window_end = _as_date(get("window_end"))
            promotion_type_raw = _cell_str(get("promotion_type_raw"))

            raw_obj = {
                str(headers[j]): (None if raw[j] is None else str(raw[j]) if not isinstance(raw[j], (int, float, date, datetime, Decimal)) else raw[j])
                for j in range(min(len(headers), len(raw)))
                if headers[j]
            }
            # JSON-safe dates
            for k, v in list(raw_obj.items()):
                if isinstance(v, datetime):
                    raw_obj[k] = v.isoformat()
                elif isinstance(v, date):
                    raw_obj[k] = v.isoformat()
                elif isinstance(v, Decimal):
                    raw_obj[k] = float(v)

            rows_out.append(
                {
                    "source_key": _source_key(
                        case_code=case_code,
                        credit_note_id=credit_note_id,
                        row_number=i,
                        sheet=sheet_name,
                    ),
                    "source_row_number": i,
                    "sheet_name": sheet_name,
                    "external_case_code": case_code,
                    "credit_note_id": credit_note_id,
                    "case_status_raw": case_status_raw,
                    "payment_status_raw": payment_status_raw,
                    "payment_status": _normalize_payment_status(payment_status_raw, status_map),
                    "payment_date": payment_date,
                    "amount": float(amount) if amount is not None else None,
                    "currency_code": currency,
                    "customer_token": customer_token,
                    "distributor_token": distributor_token,
                    "description": description,
                    "window_start": window_start,
                    "window_end": window_end,
                    "promotion_type_raw": promotion_type_raw,
                    "raw_source_row": raw_obj,
                    "flags_json": {},
                }
            )

        n = sum(1 for r in rows_out if r["sheet_name"] == sheet_name)
        sheet_summaries.append({"sheet": sheet_name, "role": "payment_evidence", "rows": n})

    wb.close()

    if not rows_out and not blocking_errors:
        blocking_errors.append(
            {
                "code": "no_payment_rows",
                "message": "No payment evidence rows parsed — check sheet roles and Case ID mapping",
            }
        )

    return {
        "ok": not blocking_errors,
        "rows": rows_out,
        "sheet_summaries": sheet_summaries,
        "profile_code": prof.get("profile_code"),
        "blocking_errors": blocking_errors,
        "row_count": len(rows_out),
    }
