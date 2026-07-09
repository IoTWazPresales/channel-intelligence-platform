"""Template-driven legacy promo plan workbook (openpyxl).

SPEC §7 / U6: This path still reads empty ``dim_promotion`` / ``fact_promotion_plan``
scaffold tables for the old ``/promotions`` export API. The authoritative CPOR XLSX
export is U4's ``/api/v1/cpor/cases/{id}/export`` (cpor_case / cpor_case_line).
Do not extend this module — re-point callers to the CPOR export instead.
"""

from __future__ import annotations

import hashlib
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.dimensions import DimCustomer, DimProduct, DimPromotion
from app.models.facts import FactPromotionPlan  # noqa: F401 — scaffold retained until table drop (spec §7)

TEMPLATE_CODE = "cpor_v1"
DATA_SHEET = "CPOR_PromoPlan"
HEADER_ROW = 1
FIRST_DATA_ROW = 2


def _build_blank_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET
    headers = [
        "Customer_Code",
        "Customer_Name",
        "Channel_Code",
        "Promo_Code",
        "Promo_Name",
        "SKU",
        "Product_Name",
        "Expected_Uplift_Pct",
        "Support_Needed",
        "Stock_Readiness",
        "Planned_Volume_Units",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col, value=h)
        cell.font = Font(bold=True)
    # Example formula row note: totals can be added in a real template file.
    ws.freeze_panes = "A2"
    return wb


def template_workbook_bytes() -> bytes:
    bio = BytesIO()
    _build_blank_template_workbook().save(bio)
    return bio.getvalue()


def validate_promotion_for_export(session: Session, promotion_id: int) -> tuple[bool, list[str], list[dict[str, Any]]]:
    errors: list[str] = []
    promo = session.get(DimPromotion, promotion_id)
    if not promo:
        return False, [f"Promotion {promotion_id} not found"], []
    rows = session.execute(select(FactPromotionPlan).where(FactPromotionPlan.promotion_id == promotion_id)).scalars().all()
    if not rows:
        errors.append("No promotion plan line items exist for this promotion.")
        return False, errors, []
    lines: list[dict[str, Any]] = []
    for r in rows:
        prod = session.get(DimProduct, r.product_id)
        if not prod:
            errors.append(f"Missing product for plan row {r.id}.")
            continue
        lines.append(
            {
                "plan_row_id": r.id,
                "sku": prod.sku,
                "product_name": prod.name,
                "expected_uplift_pct": float(r.expected_uplift_pct) if r.expected_uplift_pct is not None else None,
                "support_needed": r.support_needed,
                "stock_readiness": r.stock_readiness,
            }
        )
    if errors:
        return False, errors, []
    return True, [], lines


def build_promo_plan_workbook_bytes(
    session: Session,
    promotion_id: int,
    *,
    default_customer_id: int | None = None,
    default_channel_code: str | None = None,
) -> tuple[bytes, str]:
    ok, errors, lines = validate_promotion_for_export(session, promotion_id)
    if not ok:
        raise ValueError("; ".join(errors))

    promo = session.get(DimPromotion, promotion_id)
    assert promo is not None

    cust_code = ""
    cust_name = ""
    channel_code = default_channel_code or ""
    if default_customer_id:
        c = session.get(DimCustomer, default_customer_id)
        if c:
            cust_code = c.code
            cust_name = c.name
            if c.channel and c.channel.code:
                channel_code = c.channel.code

    bio = BytesIO(template_workbook_bytes())
    wb = load_workbook(bio)
    ws = wb[DATA_SHEET]

    row_idx = FIRST_DATA_ROW
    for line in lines:
        ws.cell(row=row_idx, column=1, value=cust_code)
        ws.cell(row=row_idx, column=2, value=cust_name)
        ws.cell(row=row_idx, column=3, value=channel_code)
        ws.cell(row=row_idx, column=4, value=promo.code)
        ws.cell(row=row_idx, column=5, value=promo.name)
        ws.cell(row=row_idx, column=6, value=line["sku"])
        ws.cell(row=row_idx, column=7, value=line["product_name"])
        ws.cell(row=row_idx, column=8, value=line["expected_uplift_pct"])
        ws.cell(row=row_idx, column=9, value=line["support_needed"])
        ws.cell(row=row_idx, column=10, value=line["stock_readiness"])
        ws.cell(row=row_idx, column=11, value="")  # optional volume planning — filled when available upstream
        row_idx += 1

    out = BytesIO()
    wb.save(out)
    data = out.getvalue()
    digest = hashlib.sha256(data).hexdigest()
    return data, digest
