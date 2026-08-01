"""P3-5 report export — Excel + minimal PDF with data vintage on the face."""

from __future__ import annotations

import io
import json
from datetime import datetime, timezone
from typing import Any


def _cell(v: Any) -> str | float | int | bool | None:
    if v is None:
        return None
    if isinstance(v, (int, float, bool, str)):
        return v
    return json.dumps(v, default=str)


def build_xlsx_bytes(
    *,
    title: str,
    metric_key: str,
    grains: list[str],
    value: Any,
    rows: list[dict[str, Any]] | None,
    data_vintage: dict[str, Any] | None,
    invariants: list[str] | None = None,
    missing_data_alert: bool = False,
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = title
    cover["A1"].font = Font(bold=True, size=14)
    cover["A2"] = f"Metric: {metric_key}"
    cover["A3"] = f"Grain: {{{', '.join(grains)}}}"
    cover["A4"] = f"Value: {_cell(value)}"
    cover["A5"] = f"Generated (UTC): {datetime.now(timezone.utc).isoformat()}"
    cover["A6"] = "Data vintage:"
    cover["A7"] = json.dumps(data_vintage or {}, default=str)
    cover["A8"] = f"Missing-data alert: {'YES — review empty/partial sources' if missing_data_alert else 'no'}"
    if invariants:
        cover["A9"] = "Invariants applied:"
        cover["A10"] = ", ".join(invariants)

    sheet = wb.create_sheet("Data")
    rows = rows or []
    if not rows:
        sheet["A1"] = "value"
        sheet["A2"] = _cell(value)
    else:
        keys = list(rows[0].keys())
        for col, k in enumerate(keys, start=1):
            sheet.cell(1, col, k).font = Font(bold=True)
        for r_i, row in enumerate(rows, start=2):
            for c_i, k in enumerate(keys, start=1):
                sheet.cell(r_i, c_i, _cell(row.get(k)))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_pdf_bytes(
    *,
    title: str,
    metric_key: str,
    grains: list[str],
    value: Any,
    rows: list[dict[str, Any]] | None,
    data_vintage: dict[str, Any] | None,
    invariants: list[str] | None = None,
    missing_data_alert: bool = False,
) -> bytes:
    """Minimal single-font PDF (no reportlab) — vintage declared on page 1."""
    lines = [
        title,
        f"Metric: {metric_key}",
        f"Grain: {{{', '.join(grains)}}}",
        f"Value: {_cell(value)}",
        f"Generated (UTC): {datetime.now(timezone.utc).isoformat()}",
        f"Data vintage: {json.dumps(data_vintage or {}, default=str)}",
        (
            "MISSING-DATA ALERT: review empty or partial sources"
            if missing_data_alert
            else "Missing-data alert: no"
        ),
    ]
    if invariants:
        lines.append("Invariants: " + ", ".join(invariants))
    lines.append("")
    lines.append("--- rows (first 40) ---")
    for row in (rows or [])[:40]:
        lines.append(json.dumps(row, default=str)[:180])
    if not rows:
        lines.append(f"(scalar) value={_cell(value)}")
    return _text_pdf(lines)


def _pdf_escape(s: str) -> str:
    return s.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _text_pdf(lines: list[str]) -> bytes:
    # Simple PDF 1.4 with one page, Helvetica, y descending from top.
    content_lines = ["BT", "/F1 10 Tf", "50 780 Td", "12 TL"]
    first = True
    for raw in lines:
        text = _pdf_escape(str(raw)[:200])
        if first:
            content_lines.append(f"({text}) Tj")
            first = False
        else:
            content_lines.append("T*")
            content_lines.append(f"({text}) Tj")
    content_lines.append("ET")
    stream = "\n".join(content_lines).encode("latin-1", errors="replace")

    objects: list[bytes] = []
    objects.append(b"1 0 obj<< /Type /Catalog /Pages 2 0 R >>endobj\n")
    objects.append(b"2 0 obj<< /Type /Pages /Kids [3 0 R] /Count 1 >>endobj\n")
    objects.append(
        b"3 0 obj<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>endobj\n"
    )
    objects.append(
        f"4 0 obj<< /Length {len(stream)} >>stream\n".encode("ascii")
        + stream
        + b"\nendstream\nendobj\n"
    )
    objects.append(b"5 0 obj<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>endobj\n")

    out = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(out))
        out.extend(obj)
    xref_pos = len(out)
    out.extend(f"xref\n0 {len(offsets)}\n".encode("ascii"))
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode("ascii"))
    out.extend(
        f"trailer<< /Size {len(offsets)} /Root 1 0 R >>\nstartxref\n{xref_pos}\n%%EOF\n".encode(
            "ascii"
        )
    )
    return bytes(out)


def detect_missing_data(result: dict[str, Any]) -> bool:
    """True when governed execute returned empty / unavailable / null value with no rows."""
    if not result.get("ok"):
        return True
    rows = result.get("rows") or []
    value = result.get("value")
    if value is None and not rows:
        return True
    vintage = result.get("data_vintage") or {}
    if vintage.get("pair_count") == 0 or vintage.get("cases_in_scope") == 0:
        return True
    return False
