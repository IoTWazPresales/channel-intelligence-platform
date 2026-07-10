"""Per-customer lineup slice export (XLSX).

Builds a customer-facing pricing slice for a single CommercialLineupCase + customer: one row per
lineup line resolved to that customer, with the full pricing chain (SRP -> ... -> calculated DAP)
sourced from the persisted ``pricing_chain_json`` and ``calc_*`` columns. Read-only; computes
nothing — it presents what the import already calculated.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.commercial_lineup import CommercialLineupCase, CommercialLineupLine
from app.models.dimensions import DimCustomer, DimDistributor, DimProduct


class LineupExportNotFoundError(LookupError):
    """Case or customer not found."""


_HEADERS = [
    "SKU",
    "Product",
    "Model",
    "Distributor",
    "Quantity",
    "New SRP (incl VAT)",
    "VAT %",
    "Dealer margin %",
    "Rebate %",
    "Disti margin %",
    "Import tax %",
    "ROE",
    "SRP ex VAT",
    "Dealer price",
    "Net price",
    "Disti cost",
    "Calculated DAP (cost ccy)",
    "Profit / unit",
    "Profit total",
    "Customer feedback",
]


def _num(v: Any) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _safe_sheet_name(label: str) -> str:
    name = re.sub(r"[\\/?*\[\]:]", "-", label).strip() or "Lineup"
    return name[:31]


async def build_customer_lineup_slice_xlsx(
    db: AsyncSession, case_id: int, customer_id: int
) -> tuple[bytes, str, int]:
    """Return (xlsx_bytes, filename, row_count) for one customer's slice of a lineup case."""
    case = await db.get(CommercialLineupCase, case_id)
    if case is None:
        raise LineupExportNotFoundError(f"Lineup case {case_id} not found")
    customer = await db.get(DimCustomer, customer_id)
    if customer is None:
        raise LineupExportNotFoundError(f"Customer {customer_id} not found")

    stmt = (
        select(
            CommercialLineupLine,
            DimProduct.sku,
            DimProduct.name,
            DimProduct.model_name,
            DimDistributor.name,
        )
        .outerjoin(DimProduct, DimProduct.id == CommercialLineupLine.product_id)
        .outerjoin(DimDistributor, DimDistributor.id == CommercialLineupLine.distributor_id)
        .where(
            CommercialLineupLine.case_id == case_id,
            CommercialLineupLine.customer_id == customer_id,
        )
        .order_by(CommercialLineupLine.source_row_number, CommercialLineupLine.id)
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name(customer.code or customer.name or "Lineup")

    # Metadata banner (kept above the table so the grid stays clean for paste/import).
    meta = [
        ("Customer", f"{customer.name or ''} ({customer.code or ''})"),
        ("Period", case.period_label or ""),
        ("Product line", case.product_line or ""),
        ("Negotiation round", case.iteration_number or 1),
        ("Currency", case.currency_code or ""),
    ]
    for i, (k, v) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)

    header_row = len(meta) + 2
    for col, h in enumerate(_HEADERS, start=1):
        ws.cell(row=header_row, column=col, value=h).font = Font(bold=True)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    row_idx = header_row + 1
    for ln, sku, pname, model, dist_name in rows:
        chain = ln.pricing_chain_json if isinstance(ln.pricing_chain_json, dict) else {}
        inp = chain.get("inputs", {}) if isinstance(chain.get("inputs"), dict) else {}
        out = chain.get("outputs", {}) if isinstance(chain.get("outputs"), dict) else {}
        values = [
            sku or ln.sku_raw,
            pname,
            model or ln.model_raw,
            dist_name,
            _num(ln.quantity_units),
            _num(inp.get("srp_inc_vat_local")) or _num(ln.msrp_local),
            _num(inp.get("vat_rate_pct")),
            _num(inp.get("dealer_margin_pct")),
            _num(inp.get("rebate_pct")),
            _num(inp.get("distributor_margin_pct")),
            _num(inp.get("import_tax_pct")),
            _num(inp.get("roe_local_per_cost_currency")),
            _num(out.get("calc_srp_ex_vat_local")),
            _num(out.get("calc_dealer_price_local")),
            _num(out.get("calc_net_price_local")),
            _num(out.get("calc_disti_cost_local")),
            _num(ln.calc_dap_cost_currency) or _num(out.get("calc_dap_cost_currency")),
            _num(out.get("calc_profit_per_unit")),
            _num(ln.calc_profit_total) or _num(out.get("calc_profit_total")),
            ln.customer_feedback or "",
        ]
        for col, val in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col, value=val)
        row_idx += 1

    bio = BytesIO()
    wb.save(bio)

    safe_code = re.sub(r"[^A-Za-z0-9_.-]+", "_", (customer.code or customer.name or str(customer_id)))
    period = re.sub(r"[^A-Za-z0-9_.-]+", "_", (case.period_label or "")) or "lineup"
    filename = f"lineup_case{case_id}_{safe_code}_{period}.xlsx"
    return bio.getvalue(), filename, len(rows)
