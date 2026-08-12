"""SCM / steward article-map import → customer_article_alias (proposed).

Column contract (synonyms accepted):
  Customer | Article code | Sales Model name

Resolves sales model → dim_product via single-match PM tiers, then stores
(customer_id, article) → product_id. Collisions (same customer+article →
multiple sales models) are reported and never auto-written. Confirmed aliases
are not overwritten. FLAG ≠ BLOCK — misses stay out of the table.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Iterable

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.customer_article_alias import CustomerArticleAlias
from app.models.dimensions import DimCustomer
from app.services.imports.cst_d1 import normalize_article_token
from app.services.imports.distributor_sales_inventory import _load_product_resolution_index
from app.services.imports.product_resolution_standard import resolve_product_id_single_match
from app.utils.json_safe import to_jsonable

# SCM display name (lower) → preferred dim_customer.name (exact, case-insensitive).
CUSTOMER_NAME_CANON: dict[str, str] = {
    "amazon": "Amazon",
    "computer mania": "Computer Mania",
    "game": "Game",
    "game online": "Game",
    "hifi": "Hifi",
    "incredible connection": "Incredible Connection",
    "makro": "Makro",
    "fnb": "FNB",
}

_CUSTOMER_HEADER = frozenset({"customer", "customer name", "customer_name", "retailer"})
_ARTICLE_HEADER = frozenset(
    {"article code", "article", "article_code", "article no", "article_no", "retailer article"}
)
_MODEL_HEADER = frozenset(
    {
        "sales model name",
        "sales model",
        "sales_model_name",
        "sales_model",
        "model",
        "model name",
    }
)


@dataclass
class AliasImportSummary:
    rows_read: int = 0
    rows_deduped: int = 0
    proposed: int = 0
    updated_proposed: int = 0
    skipped_existing_confirmed: int = 0
    collisions: int = 0
    customer_unresolved: int = 0
    model_ambiguous: int = 0
    model_miss: int = 0
    blank_skipped: int = 0
    collision_samples: list[dict[str, Any]] = field(default_factory=list)
    customer_unresolved_samples: list[str] = field(default_factory=list)
    model_miss_samples: list[str] = field(default_factory=list)
    model_ambiguous_samples: list[str] = field(default_factory=list)
    proposed_alias_ids: list[int] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _fold_header(raw: Any) -> str:
    return str(raw or "").strip().lower().replace("_", " ")


def _pick_columns(headers: Iterable[Any]) -> dict[str, str]:
    """Map logical role → original header name."""
    found: dict[str, str] = {}
    for h in headers:
        fold = _fold_header(h)
        if fold in _CUSTOMER_HEADER and "customer" not in found:
            found["customer"] = str(h)
        elif fold in _ARTICLE_HEADER and "article" not in found:
            found["article"] = str(h)
        elif fold in _MODEL_HEADER and "sales_model" not in found:
            found["sales_model"] = str(h)
    return found


def parse_article_alias_workbook(content: bytes) -> list[dict[str, str]]:
    """Parse first sheet of xlsx/xls into raw row dicts with canonical keys."""
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return []
        colmap = _pick_columns(header_row)
        if len(colmap) < 3:
            raise ValueError(
                "Articles map requires Customer, Article code, and Sales Model name columns "
                f"(found headers: {[str(h) for h in header_row if h is not None][:12]})"
            )
        out: list[dict[str, str]] = []
        for raw in rows_iter:
            by_idx = {i: raw[i] if i < len(raw) else None for i in range(len(header_row))}
            header_to_val = {str(header_row[i]): by_idx.get(i) for i in range(len(header_row))}
            cust = str(header_to_val.get(colmap["customer"]) or "").strip()
            art = str(header_to_val.get(colmap["article"]) or "").strip()
            model = str(header_to_val.get(colmap["sales_model"]) or "").strip()
            if not cust and not art and not model:
                continue
            if art.lower() == "nan":
                art = ""
            if model.lower() == "nan":
                model = ""
            out.append({"customer": cust, "article": art, "sales_model": model})
        return out
    finally:
        wb.close()


def _resolve_customer_id(session: Session, scm_name: str) -> int | None:
    raw = (scm_name or "").strip()
    if not raw:
        return None
    canon = CUSTOMER_NAME_CANON.get(raw.lower(), raw)
    row = session.scalar(
        select(DimCustomer).where(func.lower(func.btrim(DimCustomer.name)) == canon.strip().lower())
    )
    return int(row.id) if row is not None else None


def import_article_alias_rows(
    session: Session,
    rows: list[dict[str, str]],
    *,
    source: str = "scm_upload",
    actor: str | None = None,
) -> AliasImportSummary:
    """Propose aliases from parsed rows. Does not confirm. Does not overwrite confirmed."""
    summary = AliasImportSummary(rows_read=len(rows))
    # exact-row dedupe
    seen_exact: set[tuple[str, str, str]] = set()
    cleaned: list[dict[str, str]] = []
    for r in rows:
        key = (
            (r.get("customer") or "").strip().lower(),
            (r.get("article") or "").strip().lower(),
            (r.get("sales_model") or "").strip().lower(),
        )
        if key in seen_exact:
            continue
        seen_exact.add(key)
        cleaned.append(r)
    summary.rows_deduped = len(cleaned)

    idx = _load_product_resolution_index(session)
    # Group models by (customer_id, article_key)
    grouped: dict[tuple[int, str], set[str]] = defaultdict(set)
    unresolved_customers: set[str] = set()

    for r in cleaned:
        cust_raw = (r.get("customer") or "").strip()
        art_raw = (r.get("article") or "").strip()
        model_raw = (r.get("sales_model") or "").strip()
        if not cust_raw or not art_raw or not model_raw:
            summary.blank_skipped += 1
            continue
        cid = _resolve_customer_id(session, cust_raw)
        if cid is None:
            summary.customer_unresolved += 1
            unresolved_customers.add(cust_raw)
            continue
        art_key = normalize_article_token(art_raw)
        if not art_key:
            summary.blank_skipped += 1
            continue
        grouped[(cid, art_key)].add(model_raw)

    if unresolved_customers:
        summary.customer_unresolved_samples = sorted(unresolved_customers)[:20]

    now = datetime.now(timezone.utc).isoformat()
    for (cid, art_key), models in grouped.items():
        if len(models) > 1:
            summary.collisions += 1
            if len(summary.collision_samples) < 25:
                summary.collision_samples.append(
                    {
                        "customer_id": cid,
                        "article": art_key,
                        "sales_models": sorted(models),
                    }
                )
            continue

        model = next(iter(models))
        pid = resolve_product_id_single_match(idx, model)
        if pid is None:
            # Distinguish ambiguous vs miss via index
            from app.services.imports.distributor_sales_inventory import _product_token_key

            mk = _product_token_key(model)
            sm_ids = idx.sales_model_name_to_ids.get(mk) if mk else None
            if sm_ids and len(sm_ids) > 1:
                summary.model_ambiguous += 1
                if len(summary.model_ambiguous_samples) < 20:
                    summary.model_ambiguous_samples.append(model)
            else:
                summary.model_miss += 1
                if len(summary.model_miss_samples) < 20:
                    summary.model_miss_samples.append(model)
            continue

        existing = session.scalar(
            select(CustomerArticleAlias).where(
                CustomerArticleAlias.customer_id == cid,
                CustomerArticleAlias.article_no_normalized == art_key,
            )
        )
        evidence = {
            "source": source,
            "sales_model_name": model,
            "imported_at": now,
            "actor": actor,
        }
        if existing is None:
            row = CustomerArticleAlias(
                customer_id=cid,
                article_no_normalized=art_key,
                product_id=int(pid),
                status="proposed",
                evidence_json=to_jsonable(evidence),
            )
            session.add(row)
            session.flush()
            summary.proposed += 1
            summary.proposed_alias_ids.append(int(row.id))
            continue

        if existing.status in ("confirmed", "active"):
            summary.skipped_existing_confirmed += 1
            continue

        # proposed / rejected — refresh target from SCM unique match
        existing.product_id = int(pid)
        existing.status = "proposed"
        prev = dict(existing.evidence_json or {}) if isinstance(existing.evidence_json, dict) else {}
        prev.update(evidence)
        prev["refreshed_from_import"] = True
        existing.evidence_json = to_jsonable(prev)
        session.add(existing)
        summary.updated_proposed += 1
        summary.proposed_alias_ids.append(int(existing.id))

    return summary


def confirm_scm_unique_proposed(
    session: Session,
    alias_ids: list[int],
    *,
    actor: str | None = None,
) -> dict[str, int]:
    """Confirm proposed aliases that carry SCM unique-match evidence."""
    from app.services.imports.cst_d1 import confirm_customer_article_alias

    confirmed = 0
    skipped = 0
    for aid in alias_ids:
        row = session.get(CustomerArticleAlias, int(aid))
        if row is None or row.status != "proposed":
            skipped += 1
            continue
        evidence = row.evidence_json if isinstance(row.evidence_json, dict) else {}
        if evidence.get("source") not in ("scm_upload", "scm_articles_xlsx"):
            skipped += 1
            continue
        confirm_customer_article_alias(session, alias_id=int(aid), actor=actor or "scm_import")
        confirmed += 1
    return {"confirmed": confirmed, "skipped": skipped}
