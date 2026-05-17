"""Import processor: customer sell-out / POS data (retailer weekly sales reports).

Reads CSV/XLSX uploads, maps columns via ``job.field_mapping``, resolves products
through ``CustomerProductAlias`` then ``dim_product`` fallback, resolves stores via
``dim_store``, and upserts into ``fact_customer_sales``.  Unresolved product tokens
are aggregated into ``ImportEntityMappingCandidate`` rows for steward review.
"""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select, text
from sqlalchemy.orm import Session

from app.models.customer_sales import CustomerProductAlias, DimStore, FactCustomerSales
from app.models.dimensions import DimProduct
from app.models.import_distributor_si import ImportEntityMappingCandidate
from app.models.ingestion import ImportJob, ImportRowResult, RawFileMetadata
from app.storage.local import get_storage_backend
from app.utils.json_safe import to_jsonable

CUSTOMER_SALES_PRODUCT_ENTITY = "customer_sales_product"

_RE_WEEK_YEAR_1 = re.compile(r"[Ww](?:eek)?\s*(\d{1,2})\s*[,\-\s]+(\d{4})")
_RE_WEEK_YEAR_2 = re.compile(r"[Ww](\d{1,2})\s*[\-]?\s*(\d{4})")
_RE_YEAR_WEEK = re.compile(r"(\d{4})\s*[\-/][Ww]?(\d{1,2})")


def _parse_report_period(value: str) -> tuple[int | None, int | None]:
    """Parse 'Week 18 2026', 'W18-2026', or '2026-W18' into (week, year)."""
    if not value or not isinstance(value, str):
        return None, None
    v = value.strip()
    if not v:
        return None, None

    m = _RE_WEEK_YEAR_1.search(v)
    if m:
        return int(m.group(1)), int(m.group(2))

    m = _RE_WEEK_YEAR_2.search(v)
    if m:
        return int(m.group(1)), int(m.group(2))

    m = _RE_YEAR_WEEK.search(v)
    if m:
        return int(m.group(2)), int(m.group(1))

    return None, None


def _iso_week_to_date(year: int, week: int) -> date:
    """Convert ISO year+week to Monday of that week."""
    return date.fromisocalendar(year, week, 1)


def _decimal_or_none(v: Any) -> Decimal | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return Decimal(str(v).replace(",", ""))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _cell_str(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        t = v.strip()
        return t or None
    return str(v).strip() or None


def _int_or_none(v: Any) -> int | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return int(float(str(v).replace(",", "")))
    except (ValueError, TypeError):
        return None


def _parse_date(v: Any) -> date | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return None
        return ts.date()
    except (ValueError, OverflowError):
        return None


def _row_dict(series: pd.Series) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k in series.index:
        out[str(k)] = to_jsonable(series.get(k))
    return out


def _resolve_product_for_customer_sales(
    db: Session, article_code: str, customer_id: int | None
) -> tuple[int | None, str, str | None]:
    """Returns (product_id, resolution_status, resolution_detail).

    First checks ``CustomerProductAlias`` for the customer (approved, has product_id).
    Falls back to ``dim_product`` by SKU / part_number / EAN / UPC ilike match.
    """
    code = (article_code or "").strip()
    if not code:
        return None, "no_identifier", None

    if customer_id is not None:
        alias_row = db.execute(
            select(CustomerProductAlias).where(
                CustomerProductAlias.customer_id == customer_id,
                CustomerProductAlias.normalized_code == code.lower(),
                CustomerProductAlias.status == "approved",
                CustomerProductAlias.product_id.is_not(None),
            )
        ).scalars().first()
        if alias_row is not None:
            return int(alias_row.product_id), "resolved_alias", f"alias:{alias_row.id}"  # type: ignore[arg-type]

    pattern = code.strip()
    product = db.execute(
        select(DimProduct).where(
            DimProduct.is_active.is_(True),
            or_(
                DimProduct.sku.ilike(pattern),
                DimProduct.part_number.ilike(pattern),
                DimProduct.ean.ilike(pattern),
                DimProduct.upc.ilike(pattern),
            ),
        )
    ).scalars().first()
    if product is not None:
        return int(product.id), "resolved_dim", f"dim_product:{product.id}"

    return None, "no_match", None


def _resolve_store(
    db: Session, store_code: str | None, customer_id: int | None
) -> tuple[int | None, str]:
    """Returns (store_id, resolution_status)."""
    if not store_code or not isinstance(store_code, str) or not store_code.strip():
        return None, "skipped_empty"
    sc = store_code.strip()
    if customer_id is None:
        return None, "no_customer"

    store = db.execute(
        select(DimStore).where(
            DimStore.customer_id == customer_id,
            DimStore.store_code == sc,
        )
    ).scalars().first()
    if store is not None:
        return int(store.id), "resolved"
    return None, "no_match"


def _build_source_key(
    customer_id: int | None,
    year: int | None,
    week: int | None,
    article_code: str,
    store_code: str | None,
) -> str:
    """Build deterministic source key for upsert."""
    cid = str(customer_id) if customer_id is not None else "NOCUST"
    yr = str(year) if year is not None else "NOYR"
    wk = str(week) if week is not None else "NOWK"
    ac = (article_code or "").strip() or "NOART"
    sc = (store_code or "").strip() or "NOSTORE"
    return f"{cid}:{yr}:{wk}:{ac}:{sc}"


def _openpyxl_sheet_to_dataframe(ws: Any) -> pd.DataFrame:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()
    if header_row is None:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else "" for c in header_row]
    data_rows = list(rows_iter)
    if not data_rows:
        return pd.DataFrame(columns=header)
    return pd.DataFrame(list(data_rows), columns=header, dtype=object)


def _load_frames_for_job(
    job: ImportJob, df_passed: pd.DataFrame | None, raw_bytes: bytes
) -> list[tuple[str | None, pd.DataFrame]]:
    """List of (sheet_name, dataframe)."""
    fn = job.file_name or ""
    lower = fn.lower()
    out: list[tuple[str | None, pd.DataFrame]] = []

    if lower.endswith(".csv"):
        df_csv = pd.read_csv(io.BytesIO(raw_bytes))
        out.append((None, df_csv))
        return out

    if lower.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sdf = _openpyxl_sheet_to_dataframe(ws)
                out.append((sheet, sdf))
        finally:
            wb.close()
        return out

    if isinstance(df_passed, pd.DataFrame) and not df_passed.empty:
        out.append((None, df_passed))
    return out


def _invert_field_mapping(mapping: dict[str, Any] | None) -> dict[str, str]:
    """Invert field_mapping (source_header → canonical) to canonical → source_header (first wins)."""
    if not mapping:
        return {}
    rev: dict[str, str] = {}
    for src, canonical in mapping.items():
        if not isinstance(src, str) or not isinstance(canonical, str):
            continue
        s = src.strip()
        c = canonical.strip()
        if s and c and c not in rev:
            rev[c] = s
    return rev


def _get_mapped_value(row: pd.Series, canonical: str, canon_to_src: dict[str, str]) -> Any:
    """Get a cell value using the canonical→source mapping, falling back to canonical name."""
    src_header = canon_to_src.get(canonical)
    if src_header and src_header in row.index:
        v = row.get(src_header)
        if isinstance(v, pd.Series):
            return v.iloc[0] if not v.empty else None
        return v
    if canonical in row.index:
        v = row.get(canonical)
        if isinstance(v, pd.Series):
            return v.iloc[0] if not v.empty else None
        return v
    return None


def _rebuild_customer_sales_product_candidates(db: Session, job: ImportJob) -> None:
    """Build ImportEntityMappingCandidate rows for unresolved products."""
    jid = int(job.id)
    sid = int(job.source_id) if job.source_id else None

    db.execute(
        delete(ImportEntityMappingCandidate).where(
            ImportEntityMappingCandidate.import_job_id == jid,
            ImportEntityMappingCandidate.entity_type == CUSTOMER_SALES_PRODUCT_ENTITY,
        )
    )
    db.flush()

    rows = list(
        db.execute(
            select(FactCustomerSales).where(
                FactCustomerSales.import_job_id == jid,
                FactCustomerSales.product_id.is_(None),
                FactCustomerSales.product_resolution_status.in_(["no_match", "no_identifier"]),
            )
        ).scalars().all()
    )

    buckets: dict[str, dict[str, Any]] = {}
    for fact_row in rows:
        code = (fact_row.source_article_code or "").strip()
        if not code:
            continue
        nk = code.lower()[:512]
        bucket = buckets.setdefault(
            nk,
            {
                "fact_ids": [],
                "samples": [],
                "qty": Decimal(0),
                "customer_id": fact_row.customer_id,
            },
        )
        bucket["fact_ids"].append(int(fact_row.id))
        if len(bucket["samples"]) < 5 and code not in bucket["samples"]:
            bucket["samples"].append(code[:512])
        if fact_row.quantity_sold is not None:
            bucket["qty"] += Decimal(str(fact_row.quantity_sold))

    for nk, bucket in buckets.items():
        cand = ImportEntityMappingCandidate(
            import_job_id=jid,
            source_definition_id=sid,
            entity_type=CUSTOMER_SALES_PRODUCT_ENTITY,
            normalized_key=nk[:512],
            dealer_group_token=None,
            row_count=len(bucket["fact_ids"]),
            total_units=float(bucket["qty"]) if bucket["qty"] else None,
            total_reported_value=None,
            sample_raw_values=to_jsonable(bucket["samples"][:5]),
            status="needs_review",
            context=to_jsonable({
                "fact_ids": bucket["fact_ids"],
                "customer_id": bucket["customer_id"],
            }),
        )
        db.add(cand)
    db.flush()


def process_customer_sales_import(
    db: Session, job: ImportJob, df: pd.DataFrame | None = None, mapping: dict[str, str] | None = None
) -> int:
    """Process a customer sales import job. Returns error count."""
    effective_mapping: dict[str, Any] = dict(job.field_mapping or mapping or {})
    canon_to_src = _invert_field_mapping(effective_mapping)

    raw_meta = db.scalars(select(RawFileMetadata).where(RawFileMetadata.job_id == job.id)).first()
    if not raw_meta:
        db.add(
            ImportRowResult(
                job_id=job.id,
                row_number=0,
                severity="error",
                code="missing_raw_file",
                message="No raw file metadata for this import job.",
            )
        )
        return 1

    storage = get_storage_backend()
    raw_bytes = storage.read(raw_meta.storage_key)
    frames = _load_frames_for_job(job, df, raw_bytes)

    customer_id: int | None = None
    staged = job.staged_metadata or {}
    if isinstance(staged, dict):
        cid_raw = staged.get("customer_id")
        if cid_raw is not None:
            try:
                customer_id = int(cid_raw)
            except (TypeError, ValueError):
                pass

    blocking = 0
    global_row = 0

    for sheet_name, frame in frames:
        if frame is None or len(frame) == 0:
            continue

        for pos, (_, row) in enumerate(frame.iterrows(), start=2):
            global_row += 1
            try:
                series = row if isinstance(row, pd.Series) else pd.Series(row, index=frame.columns)
                raw_payload = _row_dict(series)

                source_article_code = _cell_str(
                    _get_mapped_value(series, "source_article_code", canon_to_src)
                )
                source_store_code = _cell_str(
                    _get_mapped_value(series, "source_store_code", canon_to_src)
                )

                report_period_raw = _cell_str(
                    _get_mapped_value(series, "report_period", canon_to_src)
                )
                report_week_raw = _get_mapped_value(series, "report_week", canon_to_src)
                report_year_raw = _get_mapped_value(series, "report_year", canon_to_src)
                transaction_date_raw = _get_mapped_value(series, "transaction_date", canon_to_src)

                report_week: int | None = None
                report_year: int | None = None
                transaction_date: date | None = None

                if report_period_raw:
                    report_week, report_year = _parse_report_period(report_period_raw)

                if report_week is None and report_week_raw is not None:
                    report_week = _int_or_none(report_week_raw)
                if report_year is None and report_year_raw is not None:
                    report_year = _int_or_none(report_year_raw)

                if transaction_date_raw is not None:
                    transaction_date = _parse_date(transaction_date_raw)

                if transaction_date is None and report_year is not None and report_week is not None:
                    try:
                        transaction_date = _iso_week_to_date(report_year, report_week)
                    except ValueError:
                        pass

                if report_week is None and transaction_date is not None:
                    report_week = transaction_date.isocalendar()[1]
                if report_year is None and transaction_date is not None:
                    report_year = transaction_date.isocalendar()[0]

                pid, p_status, p_detail = _resolve_product_for_customer_sales(
                    db, source_article_code or "", customer_id
                )

                sid, s_status = _resolve_store(db, source_store_code, customer_id)

                source_key = _build_source_key(
                    customer_id, report_year, report_week, source_article_code or "", source_store_code
                )

                quantity_sold = _decimal_or_none(
                    _get_mapped_value(series, "quantity_sold", canon_to_src)
                )
                quantity_returned = _decimal_or_none(
                    _get_mapped_value(series, "quantity_returned", canon_to_src)
                )
                selling_price = _decimal_or_none(
                    _get_mapped_value(series, "selling_price", canon_to_src)
                )
                cost_price = _decimal_or_none(
                    _get_mapped_value(series, "cost_price", canon_to_src)
                )
                currency_code = _cell_str(
                    _get_mapped_value(series, "currency_code", canon_to_src)
                )
                channel_type = _cell_str(
                    _get_mapped_value(series, "channel_type", canon_to_src)
                )
                reported_soh = _decimal_or_none(
                    _get_mapped_value(series, "reported_soh", canon_to_src)
                )

                params = {
                    "source_key": source_key[:256],
                    "customer_id": customer_id,
                    "product_id": pid,
                    "store_id": sid,
                    "import_job_id": int(job.id),
                    "report_week": report_week,
                    "report_year": report_year,
                    "report_period": (report_period_raw or "")[:32] or None,
                    "transaction_date": transaction_date,
                    "quantity_sold": float(quantity_sold) if quantity_sold is not None else None,
                    "quantity_returned": float(quantity_returned) if quantity_returned is not None else None,
                    "selling_price": float(selling_price) if selling_price is not None else None,
                    "cost_price": float(cost_price) if cost_price is not None else None,
                    "currency_code": (currency_code or "")[:8] or None,
                    "channel_type": (channel_type or "")[:32] or None,
                    "reported_soh": float(reported_soh) if reported_soh is not None else None,
                    "source_article_code": (source_article_code or "")[:512] or None,
                    "source_store_code": (source_store_code or "")[:128] or None,
                    "product_resolution_status": p_status,
                    "store_resolution_status": s_status,
                    "raw_source_row": to_jsonable(raw_payload),
                }

                with db.begin_nested():
                    db.execute(text("""
                        INSERT INTO fact_customer_sales (source_key, customer_id, product_id, store_id, import_job_id,
                            report_week, report_year, report_period, transaction_date,
                            quantity_sold, quantity_returned, selling_price, cost_price, currency_code,
                            channel_type, reported_soh, source_article_code, source_store_code,
                            product_resolution_status, store_resolution_status, raw_source_row)
                        VALUES (:source_key, :customer_id, :product_id, :store_id, :import_job_id,
                            :report_week, :report_year, :report_period, :transaction_date,
                            :quantity_sold, :quantity_returned, :selling_price, :cost_price, :currency_code,
                            :channel_type, :reported_soh, :source_article_code, :source_store_code,
                            :product_resolution_status, :store_resolution_status, :raw_source_row)
                        ON CONFLICT (source_key) DO UPDATE SET
                            product_id = EXCLUDED.product_id,
                            store_id = EXCLUDED.store_id,
                            product_resolution_status = EXCLUDED.product_resolution_status,
                            store_resolution_status = EXCLUDED.store_resolution_status,
                            raw_source_row = EXCLUDED.raw_source_row,
                            import_job_id = EXCLUDED.import_job_id,
                            updated_at = now()
                    """), params)

            except Exception as exc:  # noqa: BLE001
                blocking += 1
                db.add(
                    ImportRowResult(
                        job_id=job.id,
                        row_number=global_row,
                        severity="error",
                        code="customer_sales_row_error",
                        message=str(exc)[:2000],
                        raw_payload={"sheet": sheet_name, "row_index": pos},
                    )
                )

    db.flush()

    if global_row == 0:
        blocking += 1
        db.add(
            ImportRowResult(
                job_id=job.id,
                row_number=0,
                severity="error",
                code="customer_sales_empty_file",
                message="No data rows found in uploaded file.",
            )
        )

    db.flush()
    _rebuild_customer_sales_product_candidates(db, job)

    meta = dict(job.staged_metadata or {})
    meta["customer_sales"] = to_jsonable({
        "processed_at": datetime.now(timezone.utc).isoformat(),
        "total_rows": global_row,
        "blocking_errors": blocking,
    })
    job.staged_metadata = to_jsonable(meta)

    summary = {
        "rows": global_row,
        "blocking": blocking,
    }
    db.add(
        ImportRowResult(
            job_id=job.id,
            row_number=0,
            severity="info" if blocking == 0 else "warning",
            code="customer_sales_summary",
            message=json.dumps(summary),
        )
    )
    return 1 if blocking else 0
