"""Import processor: shipment / order evidence (XXOMRPT0025/0027, ACZA-style workbooks).

Auto-detects report shape from headers. Perserves each source row in ``raw_source_row``.
Product resolution: strongest token first (Item → EAN → UPC → sales model), reusing DSI
``_resolve_product`` / ``ProductResolutionIndex``. Stops at first **ambiguous** outcome.
Distributor: Bill To, then Ship To via ``_resolve_distributor_strict`` (alias + exact dim only).
"""

from __future__ import annotations

import io
import json
import re
from collections import defaultdict
from datetime import date, datetime, timezone
from decimal import Decimal
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from psycopg.errors import DataError as PsycopgDataError
from sqlalchemy import delete, func, or_, select
from sqlalchemy.exc import DBAPIError
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models.import_distributor_si import ImportEntityMappingCandidate
from app.models.ingestion import ImportJob, ImportRowResult, RawFileMetadata, SourceDefinition
from app.models.shipment_evidence import ShipmentEvidenceLine
from app.services.imports.distributor_sales_inventory import (
    DSIResolutionCache,
    ProductResolutionIndex,
    _build_distributor_resolution_cache,
    _load_product_resolution_index,
    _norm_key,
    _resolve_distributor_strict,
    _resolve_distributor_strict_from_cache,
    _resolve_product,
)
from app.services.imports.shipment_evidence_candidate_names import suggested_name_for_distributor_token
from app.services.imports.shipment_evidence_customer_remainder_merge import (
    apply_intra_job_remainder_merge_pass,
)
from app.services.imports.shipment_evidence_customer_token_naming import (
    CustomerTokenNamingResult,
    annotate_shipment_customer_pending_duplicates,
    detect_statistical_prefixes,
    grouped_candidate_normalized_key,
    plural_merge_canonical_display,
    suggest_customer_token_name,
)
from app.services.imports.shipment_evidence_resolution_plan import (
    SHIPMENT_CUSTOMER_ENTITY,
    SHIPMENT_DISTRIBUTOR_ENTITY,
    enrich_shipment_customer_token_candidates,
    enrich_shipment_distributor_candidates,
)
from app.services.imports.shipment_evidence_source_keys import (
    ShipmentEvidenceSourceKeyError,
    stable_source_key_for_row,
)
from app.services.imports.shipment_evidence_report_detect import (
    LINE_OPEN_ORDER,
    LINE_SHIPPED,
    REPORT_ACZA_SHIPPED,
    REPORT_ACZA_UNSHIP,
    REPORT_UNKNOWN,
    REPORT_XXOMRPT0025,
    REPORT_XXOMRPT0027,
    _ean_upc_str,
    detect_report_type,
)
from app.services.imports.shipment_evidence_text_normalize import normalize_shipment_cell_value
from app.storage.local import get_storage_backend
from app.utils.json_safe import to_jsonable

# Excel 9999-12-31 serial; above this is not a representable calendar date in Excel date space.
_EXCEL_MAX_DATE_SERIAL = 2958465
_RE_EXTREME_LEADING_YEAR_DATE_STR = re.compile(r"^\d{5,}-")


def _norm_cols(cols: list[str]) -> set[str]:
    return {str(c).strip() for c in cols if c is not None}


def _cell_str(v: Any) -> str | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        t = v.strip()
        return t or None
    if hasattr(v, "isoformat"):
        return None
    return str(v).strip() or None


def _parse_date(v: Any) -> date | None:
    """Parse a cell to a date safe for PostgreSQL ``DATE``; None if invalid or out of range."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, str):
        st = v.strip()
        if not st:
            return None
        if _RE_EXTREME_LEADING_YEAR_DATE_STR.match(st):
            return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and pd.isna(v):
            return None
        try:
            fv = float(v)
        except (TypeError, ValueError):
            pass
        else:
            if fv > float(_EXCEL_MAX_DATE_SERIAL):
                return None
    if isinstance(v, datetime):
        d = v.date()
        if d.year < 1 or d.year > 9999:
            return None
        return d
    if isinstance(v, date):
        if v.year < 1 or v.year > 9999:
            return None
        return v
    try:
        ts = pd.to_datetime(v, errors="coerce")
        if pd.isna(ts):
            return None
        if not isinstance(ts, pd.Timestamp):
            return None
        try:
            d = ts.date()
        except (OverflowError, ValueError, OSError):
            return None
        if d.year < 1 or d.year > 9999:
            return None
        return d
    except (ValueError, OverflowError, PsycopgDataError):
        return None


def _row_dict(series: pd.Series) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for k in series.index:
        out[str(k)] = to_jsonable(series.get(k))
    return out


def _normalized_identifier_cell(v: Any) -> Any:
    """Prefer normalized string; fall back to original cell for numeric EAN/UPC cells."""
    nu = normalize_shipment_cell_value(v)
    return nu if nu is not None else v


def _tabular_cell_scalar(v: Any) -> Any:
    """Coerce pandas duplicate-label cell access to a scalar (truth tests and ``or`` chains are safe)."""
    if isinstance(v, pd.Series):
        if v.empty:
            return None
        return v.iloc[0]
    return v


def _extract_common(row: pd.Series, header_by_canonical: dict[str, str] | None = None) -> dict[str, Any]:
    """Extract canonical shipment fields. When ``header_by_canonical`` is set (canonical -> file header), use it first."""

    def by_canon(canon: str) -> Any:
        if header_by_canonical:
            hdr = header_by_canonical.get(canon)
            if hdr and hdr in row.index:
                return _tabular_cell_scalar(row.get(hdr))
        return None

    def col(*names: str) -> Any:
        for n in names:
            if n in row.index:
                return _tabular_cell_scalar(row.get(n))
        return None

    has_distributor_token_header = bool(header_by_canonical and header_by_canonical.get("distributor_token"))
    dist_cell = by_canon("distributor_token")
    bill = by_canon("bill_to_raw") or col("Bill To", "bill to")
    if has_distributor_token_header:
        dist_norm = normalize_shipment_cell_value(dist_cell)
        bill_norm = normalize_shipment_cell_value(bill)
        bill_to_stored = dist_norm if dist_norm else bill_norm
    else:
        bill_to_stored = normalize_shipment_cell_value(bill)
    ship = by_canon("ship_to_raw") or col("Ship To", "ship to")
    ou = by_canon("operating_unit") or col("Operating Unit", "OU NAME", "ou name")
    return {
        "operating_unit": normalize_shipment_cell_value(ou),
        "bill_to_raw": bill_to_stored,
        "ship_to_raw": normalize_shipment_cell_value(ship),
        "order_no": normalize_shipment_cell_value(
            by_canon("order_no")
            or col("Order No.", "Order No", "order no.", "order no", "ORDER NO", "Order number")
        ),
        "customer_po": normalize_shipment_cell_value(
            by_canon("customer_po")
            or col(
                "Customer PO",
                "Cust PO",
                "Customer P/O",
                "Purchase Order",
                "PO No",
                "PO No.",
                "PO Number",
            )
        ),
        "order_line": normalize_shipment_cell_value(by_canon("order_line") or col("Order Line", "order line")),
        "delivery_no": normalize_shipment_cell_value(
            by_canon("delivery_no")
            or col("Delivery No", "delivery no", "DELIVERY NO", "Delivery no.", "Delivery Number", "delivery number")
        ),
        "invoice_line": normalize_shipment_cell_value(by_canon("invoice_line") or col("Invoice Line")),
        "item_code": normalize_shipment_cell_value(by_canon("item_code") or col("Item")),
        "sales_model_name": normalize_shipment_cell_value(by_canon("sales_model_name") or col("Sales Model Name")),
        "customer_item": normalize_shipment_cell_value(by_canon("customer_item") or col("Customer Item")),
        "ean_code": _ean_upc_str(
            _normalized_identifier_cell(by_canon("ean_code") or col("EAN Code"))
        ),
        "upc_code": _ean_upc_str(
            _normalized_identifier_cell(by_canon("upc_code") or col("UPC Code"))
        ),
        "mpor_item_no": normalize_shipment_cell_value(by_canon("mpor_item_no") or col("MPOR Item No.")),
        "quantity": by_canon("quantity")
        if by_canon("quantity") is not None
        else _tabular_cell_scalar(row.get("Qty") if "Qty" in row.index else row.get("Qty ")),
        "unit_price": by_canon("unit_price")
        if by_canon("unit_price") is not None
        else _tabular_cell_scalar(row.get("Unit Price") if "Unit Price" in row.index else None),
        "amount": by_canon("amount")
        if by_canon("amount") is not None
        else _tabular_cell_scalar(row.get("Amount") if "Amount" in row.index else None),
        "currency_code": normalize_shipment_cell_value(by_canon("currency_code") or col("Currency")),
        "ship_confirm_date": _parse_date(by_canon("ship_confirm_date") or col("Ship Confirm Date")),
        "schedule_ship_date": _parse_date(by_canon("schedule_ship_date") or col("Schedule Ship Date")),
        "promise_date": _parse_date(by_canon("promise_date") or col("Promise Date")),
        "exwork_date": _parse_date(by_canon("exwork_date") or col("Exwork Date")),
        "erd_date": _parse_date(by_canon("erd_date") or col("ERD (Est Revenue Date)")),
        "est_pod_date": _parse_date(
            by_canon("est_pod_date")
            or col(
                "Est POD Date",
                "Estimated POD Date",
                "Estimated Proof of Delivery",
                "Expected Delivery Date",
            )
        ),
        "pod_date": _parse_date(
            by_canon("pod_date")
            or col(
                "POD Date",
                "Proof of Delivery",
                "Actual Delivery Date",
                "Delivery Confirmed Date",
            )
        ),
        "customer_dealer_token": normalize_shipment_cell_value(
            by_canon("customer_dealer_token")
            or col(
                "Customer Remarks",
                "customer remarks",
                "Customer Remark",
                "customer remark",
                "CUSTOMER REMARKS",
            )
        ),
    }


def _decimal_or_none(v: Any) -> Decimal | None:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    try:
        return Decimal(str(v).replace(",", ""))
    except Exception:
        return None


def coalesce_shipment_evidence_date(
    *,
    pod_date: date | None = None,
    est_pod_date: date | None = None,
    promise_date: date | None = None,
    schedule_ship_date: date | None = None,
    ship_confirm_date: date | None = None,
    erd_date: date | None = None,
    exwork_date: date | None = None,
) -> date | None:
    """First non-null logistics date (pod-first), matching inbound fact apply path."""
    for d in (
        pod_date,
        est_pod_date,
        promise_date,
        schedule_ship_date,
        ship_confirm_date,
        erd_date,
        exwork_date,
    ):
        if d is not None:
            return d
    return None


def resolve_product_for_evidence(
    idx: ProductResolutionIndex,
    *,
    item_code: str | None,
    ean_code: str | None,
    upc_code: str | None,
    sales_model_name: str | None,
    evidence_date: date | None = None,
) -> tuple[int | None, str, str | None, str | None]:
    """(product_id, status, token_used, detail)."""
    tokens: list[tuple[str, str]] = []
    if item_code:
        tokens.append(("item", item_code))
    if ean_code:
        tokens.append(("ean", ean_code))
    if upc_code:
        tokens.append(("upc", upc_code))
    if sales_model_name:
        tokens.append(("sales_model", sales_model_name))
    if not tokens:
        return None, "no_identifier", None, None

    last_detail: str | None = None
    for role, raw in tokens:
        pid, perr, tag, ev = _resolve_product(
            raw,
            idx,
            evidence_date,
            shipment_sku_item_code_anchor=(role == "item"),
        )
        if pid is not None and perr is None:
            return int(pid), "resolved_unique", raw, tag or "resolved"
        if perr in ("ambiguous_product_match", "ambiguous_product_alias"):
            return None, "ambiguous", raw, perr
        if ev is not None and getattr(ev, "ambiguous_eligible", None):
            return None, "ambiguous", raw, "ambiguous_eligible"
        if perr == "unresolved_product_inactive_only":
            last_detail = perr
            continue
        last_detail = perr or tag or "unresolved"
    if last_detail == "unresolved_product_inactive_only":
        return None, "inactive_only", tokens[0][1], last_detail
    return None, "no_match", tokens[-1][1], last_detail


def resolve_distributor_for_evidence(
    db: Session,
    source_id: int | None,
    *,
    bill_to: str | None,
    ship_to: str | None,
    res_cache: "DSIResolutionCache | None" = None,
) -> tuple[int | None, str, str | None]:
    """Resolve a line's distributor via Bill To then Ship To.

    When ``res_cache`` is supplied, resolution is fully in-memory (zero per-row DB queries) —
    the strict alias + exact code/name semantics are identical to the per-row DB path.
    """

    def _resolve_token(tok: str) -> int | None:
        if res_cache is not None:
            did, _err = _resolve_distributor_strict_from_cache(tok, source_id, res_cache)
        else:
            did, _err = _resolve_distributor_strict(db, tok, source_id)
        return int(did) if did is not None else None

    if bill_to:
        did = _resolve_token(bill_to)
        if did is not None:
            return did, "resolved", bill_to
    if ship_to:
        did = _resolve_token(ship_to)
        if did is not None:
            return did, "resolved", ship_to
    if not (bill_to or ship_to):
        return None, "skipped_empty", None
    return None, "unresolved", bill_to or ship_to


def _rebuild_shipment_distributor_candidates(db: Session, job: ImportJob) -> None:
    """Replace ``shipment_distributor`` candidates for this job from unresolved evidence lines."""
    jid = int(job.id)
    sid = int(job.source_id) if job.source_id else None

    db.execute(
        delete(ImportEntityMappingCandidate).where(
            ImportEntityMappingCandidate.import_job_id == jid,
            ImportEntityMappingCandidate.entity_type == SHIPMENT_DISTRIBUTOR_ENTITY,
        )
    )
    db.flush()

    lines = list(db.scalars(select(ShipmentEvidenceLine).where(ShipmentEvidenceLine.import_job_id == jid)).all())
    buckets: dict[str, dict[str, Any]] = {}

    for line in lines:
        if line.distributor_id is not None:
            continue
        if (line.distributor_resolution_status or "") != "unresolved":
            continue
        raw: str | None = None
        party: str | None = None
        btr = line.bill_to_raw
        strw = line.ship_to_raw
        if btr and str(btr).strip():
            raw = str(btr).strip()
            party = "bill_to"
        elif strw and str(strw).strip():
            raw = str(strw).strip()
            party = "ship_to"
        else:
            continue
        nk = _norm_key(raw)
        if not nk:
            continue
        bucket = buckets.setdefault(
            nk,
            {"line_ids": [], "samples": [], "parties": set(), "qty": Decimal(0), "amt": Decimal(0)},
        )
        bucket["line_ids"].append(int(line.id))
        if party:
            bucket["parties"].add(party)
        if len(bucket["samples"]) < 5 and raw not in bucket["samples"]:
            bucket["samples"].append(raw[:512])
        if line.quantity is not None:
            bucket["qty"] += Decimal(str(line.quantity))
        if line.amount is not None:
            bucket["amt"] += Decimal(str(line.amount))

    for nk, bucket in buckets.items():
        primary_party = "bill_to" if "bill_to" in bucket["parties"] else "ship_to"
        sample0 = bucket["samples"][0] if bucket["samples"] else nk
        dis_sug = suggested_name_for_distributor_token(str(sample0))
        cand = ImportEntityMappingCandidate(
            import_job_id=jid,
            source_definition_id=sid,
            entity_type=SHIPMENT_DISTRIBUTOR_ENTITY,
            normalized_key=nk[:512],
            dealer_group_token=None,
            row_count=len(bucket["line_ids"]),
            total_units=float(bucket["qty"]) if bucket["qty"] else None,
            total_reported_value=float(bucket["amt"]) if bucket["amt"] else None,
            sample_raw_values=to_jsonable(bucket["samples"][:5]),
            status="needs_review",
            context=to_jsonable(
                {"party": primary_party, "line_ids": bucket["line_ids"], "suggested_name": dis_sug}
            ),
        )
        db.add(cand)
    db.flush()
    enrich_shipment_distributor_candidates(db, import_job_id=jid, source_definition_id=sid)


def _shipment_customer_candidate_bucket_key(nr: CustomerTokenNamingResult) -> tuple[str, str | None]:
    """Merge lines that share the same steward outcome (post all naming layers).

    Non-noise: bucket by normalised suggested display so two strings that ``_norm_key``
    the same cannot produce duplicate ``normalized_key`` rows. ``noise_only`` /
    ``internal_note``: keep buckets keyed by the per-row suggested display plus category.
    """
    if nr.special_category == "noise_only":
        return (nr.suggested_name, nr.special_category)
    if nr.special_category == "internal_note":
        return (nr.suggested_name, nr.special_category)
    nsk = _norm_key(nr.suggested_name)
    if nsk:
        return (nsk, nr.special_category)
    return ((nr.suggested_name or "").strip().lower(), nr.special_category)


def _merge_pending_plural_customer_groups(pending: dict[str, dict[str, Any]]) -> None:
    """Merge pending buckets whose displays differ only by a trailing plural ``s`` (guard: stem >= 4 chars)."""
    sig_to_nks: dict[tuple[Any, ...], list[str]] = defaultdict(list)
    for nk, pb in pending.items():
        sug = (pb.get("display_suggested_name") or "").strip()
        sc = pb.get("special_category")
        if sc in ("noise_only", "internal_note"):
            sig_to_nks[(nk, sc)].append(nk)
        else:
            canon = plural_merge_canonical_display(sug)
            sig_to_nks[(_norm_key(canon), sc)].append(nk)

    new_pending: dict[str, dict[str, Any]] = {}
    for _sig, nk_list in sig_to_nks.items():
        merged: dict[str, Any] | None = None
        for nk in nk_list:
            pb = pending.pop(nk)
            if merged is None:
                merged = {
                    "line_ids": list(pb["line_ids"]),
                    "samples": list(pb.get("samples", [])[:5]),
                    "source_tokens": list(pb.get("source_tokens", [])),
                    "qty": pb["qty"],
                    "amt": pb["amt"],
                    "needs_name_review": bool(pb.get("needs_name_review")),
                    "display_suggested_name": (pb.get("display_suggested_name") or "").strip()[:256] or None,
                    "special_category": pb.get("special_category"),
                }
            else:
                merged["line_ids"].extend(pb["line_ids"])
                merged["source_tokens"] = sorted(set(merged["source_tokens"]) | set(pb.get("source_tokens", [])))
                merged["qty"] += pb["qty"]
                merged["amt"] += pb["amt"]
                merged["needs_name_review"] = bool(merged["needs_name_review"] or pb.get("needs_name_review"))
                for s in pb.get("samples", []):
                    if len(merged["samples"]) < 5 and s not in merged["samples"]:
                        merged["samples"].append(s)
                d2 = (pb.get("display_suggested_name") or "").strip()
                if d2:
                    cur = (merged.get("display_suggested_name") or "").strip()
                    c2 = plural_merge_canonical_display(d2)
                    if not cur or len(c2) < len(cur):
                        merged["display_suggested_name"] = c2[:256]
        if merged is None:
            continue
        sc = merged.get("special_category")
        src_sorted = sorted(set(merged["source_tokens"]))
        disp = (merged.get("display_suggested_name") or "").strip()
        if sc not in ("noise_only", "internal_note") and disp:
            disp = plural_merge_canonical_display(disp)[:256]
        nk_new = grouped_candidate_normalized_key(
            suggested_name=disp or (src_sorted[0] if src_sorted else ""),
            source_tokens=src_sorted,
            special_category=sc,
        )[:512]
        merged["display_suggested_name"] = disp[:256] if disp else merged.get("display_suggested_name")
        merged["source_tokens"] = src_sorted
        new_pending[nk_new] = merged

    pending.clear()
    pending.update(new_pending)


def _distributor_suggested_display_names_for_job(db: Session, import_job_id: int) -> frozenset[str]:
    """Suggested distributor display strings already persisted for this import job (customer merge hints)."""
    rows = list(
        db.scalars(
            select(ImportEntityMappingCandidate).where(
                ImportEntityMappingCandidate.import_job_id == int(import_job_id),
                ImportEntityMappingCandidate.entity_type == SHIPMENT_DISTRIBUTOR_ENTITY,
            )
        ).all()
    )
    out: set[str] = set()
    for row in rows:
        ctx = row.context if isinstance(row.context, dict) else {}
        sn = ctx.get("suggested_name")
        if isinstance(sn, str) and sn.strip():
            out.add(sn.strip()[:256])
    return frozenset(out)


def _rebuild_shipment_customer_candidates(db: Session, job: ImportJob) -> None:
    """Replace ``shipment_customer_token`` candidates from lines with a customer token that are not steward-resolved."""
    jid = int(job.id)
    sid = int(job.source_id) if job.source_id else None

    db.execute(
        delete(ImportEntityMappingCandidate).where(
            ImportEntityMappingCandidate.import_job_id == jid,
            ImportEntityMappingCandidate.entity_type == SHIPMENT_CUSTOMER_ENTITY,
        )
    )
    db.flush()

    lines = list(db.scalars(select(ShipmentEvidenceLine).where(ShipmentEvidenceLine.import_job_id == jid)).all())
    line_refs: list[tuple[ShipmentEvidenceLine, str]] = []
    for line in lines:
        if (line.customer_resolution_status or "").strip() == "resolved":
            continue
        ctr = line.customer_dealer_token
        if not ctr or not str(ctr).strip():
            continue
        raw = str(ctr).strip()
        if not _norm_key(raw):
            continue
        line_refs.append((line, raw))

    if not line_refs:
        db.flush()
        return

    distinct_raws = sorted({r for _, r in line_refs})
    stat_prefixes, prefix_job_meta = detect_statistical_prefixes(distinct_raws)
    source_def = db.get(SourceDefinition, int(sid)) if sid else None

    naming_by_raw: dict[str, CustomerTokenNamingResult] = {}
    for r in distinct_raws:
        naming_by_raw[r] = suggest_customer_token_name(
            r, statistical_prefixes_longest_first=stat_prefixes, source_def=source_def
        )

    # Group by post-layer outcome: non-noise merges on normalised suggested_name so
    # ``grouped_candidate_normalized_key`` cannot collide across buckets.
    groups: dict[tuple[str, str | None], dict[str, Any]] = defaultdict(
        lambda: {
            "line_ids": [],
            "samples": [],
            "source_tokens": set(),
            "qty": Decimal(0),
            "amt": Decimal(0),
            "needs_name_review": False,
            "display_suggested_name": None,
        }
    )
    for line, raw in line_refs:
        nr = naming_by_raw[raw]
        gk = _shipment_customer_candidate_bucket_key(nr)
        b = groups[gk]
        if b["display_suggested_name"] is None and (nr.suggested_name or "").strip():
            b["display_suggested_name"] = (nr.suggested_name or "").strip()[:256]
        b["line_ids"].append(int(line.id))
        b["source_tokens"].add(raw[:512])
        if len(b["samples"]) < 5 and raw[:512] not in b["samples"]:
            b["samples"].append(raw[:512])
        if line.quantity is not None:
            b["qty"] += Decimal(str(line.quantity))
        if line.amount is not None:
            b["amt"] += Decimal(str(line.amount))
        if nr.needs_name_review:
            b["needs_name_review"] = True

    # Second pass: merge any buckets that still map to the same normalized_key (safety net).
    pending: dict[str, dict[str, Any]] = {}
    for merge_key, bucket in groups.items():
        special_cat = merge_key[1]
        sug_display = (bucket.get("display_suggested_name") or "").strip()
        if not sug_display:
            sug_display = merge_key[0] if special_cat == "noise_only" else (merge_key[0] or "").strip()
        src_sorted = sorted(bucket["source_tokens"])
        nk = grouped_candidate_normalized_key(
            suggested_name=sug_display,
            source_tokens=src_sorted,
            special_category=special_cat,
        )
        nk = nk[:512]
        if nk in pending:
            p = pending[nk]
            p["line_ids"].extend(bucket["line_ids"])
            p["source_tokens"] = sorted(set(p["source_tokens"]) | set(src_sorted))
            p["qty"] += bucket["qty"]
            p["amt"] += bucket["amt"]
            p["needs_name_review"] = bool(p["needs_name_review"] or bucket["needs_name_review"])
            for s in bucket["samples"]:
                if len(p["samples"]) < 5 and s not in p["samples"]:
                    p["samples"].append(s)
            if not (p.get("display_suggested_name") or "").strip() and sug_display:
                p["display_suggested_name"] = sug_display[:256]
        else:
            pending[nk] = {
                "line_ids": list(bucket["line_ids"]),
                "samples": list(bucket["samples"][:5]),
                "source_tokens": list(src_sorted),
                "qty": bucket["qty"],
                "amt": bucket["amt"],
                "needs_name_review": bool(bucket["needs_name_review"]),
                "display_suggested_name": sug_display[:256] if sug_display else None,
                "special_category": special_cat,
            }

    _merge_pending_plural_customer_groups(pending)
    annotate_shipment_customer_pending_duplicates(pending)
    dist_names = _distributor_suggested_display_names_for_job(db, jid)
    apply_intra_job_remainder_merge_pass(pending, distributor_suggested_names=dist_names)

    for nk, pb in pending.items():
        src_sorted = sorted(set(pb["source_tokens"]))
        sug_out = (pb.get("display_suggested_name") or "").strip()
        if not sug_out:
            if nk.startswith("sc:") or nk.startswith("blank:") or nk.startswith("in:"):
                sug_out = (src_sorted[0] if src_sorted else "")[:256]
            else:
                sug_out = nk[:256]
        ctx: dict[str, Any] = {
            "line_ids": pb["line_ids"],
            "suggested_name": sug_out,
            "source_tokens": src_sorted,
            "needs_name_review": bool(pb["needs_name_review"]),
            "statistical_prefixes_detected": prefix_job_meta,
        }
        sc = pb.get("special_category")
        if sc:
            ctx["special_category"] = sc
        dup = pb.get("possible_duplicate_of")
        if isinstance(dup, list) and dup:
            ctx["possible_duplicate_of"] = dup
        typo = pb.get("typo_suspected_of")
        if isinstance(typo, list) and typo:
            ctx["typo_suspected_of"] = typo
        cand = ImportEntityMappingCandidate(
            import_job_id=jid,
            source_definition_id=sid,
            entity_type=SHIPMENT_CUSTOMER_ENTITY,
            normalized_key=nk[:512],
            dealer_group_token=None,
            row_count=len(pb["line_ids"]),
            total_units=float(pb["qty"]) if pb["qty"] else None,
            total_reported_value=float(pb["amt"]) if pb["amt"] else None,
            sample_raw_values=to_jsonable(pb["samples"][:5]),
            status="needs_review",
            context=to_jsonable(ctx),
        )
        db.add(cand)
    db.flush()
    enrich_shipment_customer_token_candidates(db, import_job_id=jid, source_definition_id=sid)


# Multi-row upsert batch size. ~41 columns/row keeps params well under Postgres' 65535 limit
# (1000 × 41 ≈ 41k) while cutting round-trips ~5× vs small batches against a remote DB.
_SHIPMENT_UPSERT_CHUNK = 1000

_SHIPMENT_LINE_DATE_COLS = (
    "ship_confirm_date",
    "schedule_ship_date",
    "promise_date",
    "exwork_date",
    "erd_date",
    "est_pod_date",
    "pod_date",
)


def _is_psycopg_data_error(exc: BaseException) -> bool:
    if isinstance(exc, PsycopgDataError):
        return True
    return isinstance(exc, DBAPIError) and isinstance(getattr(exc, "orig", None), PsycopgDataError)


def _shipment_line_conflict_set(ex: Any) -> dict[str, Any]:
    """ON CONFLICT DO UPDATE assignments (source-derived columns only).

    ``id``, ``created_at`` and all product/distributor/customer resolution columns on the
    existing row are intentionally preserved; post-loop ``_resolve_unresolved_shipment_lines_for_job``
    fills unresolved ids. Shared by the single-row and bulk upsert paths.
    """
    return {
        "source_sheet": ex.source_sheet,
        "source_row_number": ex.source_row_number,
        "report_type": ex.report_type,
        "line_state": ex.line_state,
        "raw_source_row": ex.raw_source_row,
        "operating_unit": ex.operating_unit,
        "bill_to_raw": ex.bill_to_raw,
        "ship_to_raw": ex.ship_to_raw,
        "order_no": ex.order_no,
        "customer_po": ex.customer_po,
        "order_line": ex.order_line,
        "delivery_no": ex.delivery_no,
        "invoice_line": ex.invoice_line,
        "item_code": ex.item_code,
        "sales_model_name": ex.sales_model_name,
        "customer_item": ex.customer_item,
        "ean_code": ex.ean_code,
        "upc_code": ex.upc_code,
        "mpor_item_no": ex.mpor_item_no,
        "quantity": ex.quantity,
        "unit_price": ex.unit_price,
        "amount": ex.amount,
        "currency_code": ex.currency_code,
        "ship_confirm_date": ex.ship_confirm_date,
        "schedule_ship_date": ex.schedule_ship_date,
        "promise_date": ex.promise_date,
        "exwork_date": ex.exwork_date,
        "erd_date": ex.erd_date,
        "est_pod_date": ex.est_pod_date,
        "pod_date": ex.pod_date,
        "customer_dealer_token": ex.customer_dealer_token,
        "updated_at": func.now(),
    }


def _shipment_evidence_line_upsert_statement(values: dict[str, Any]):
    t = ShipmentEvidenceLine.__table__
    ins = pg_insert(t).values(**values)
    return ins.on_conflict_do_update(
        constraint="uq_shipment_evidence_line_import_job_source_key",
        set_=_shipment_line_conflict_set(ins.excluded),
    )


def _shipment_evidence_line_bulk_upsert_statement(rows: list[dict[str, Any]]):
    t = ShipmentEvidenceLine.__table__
    ins = pg_insert(t).values(rows)
    return ins.on_conflict_do_update(
        constraint="uq_shipment_evidence_line_import_job_source_key",
        set_=_shipment_line_conflict_set(ins.excluded),
    )


def _execute_shipment_line_upsert(db: Session, values: dict[str, Any]) -> None:
    """Insert or update one line keyed by (import_job_id, source_key).

    On conflict, refreshes source-derived columns only; ``id``, ``created_at``, and all
    product/distributor/customer resolution columns on the existing row are left unchanged (see
    post-loop ``_resolve_unresolved_shipment_lines_for_job`` for unresolved ids).

    If PostgreSQL rejects a bound date (e.g. year > 9999 from Excel artifacts), clears
    date columns and retries once inside a savepoint so the outer import transaction survives.
    """
    stmt = _shipment_evidence_line_upsert_statement(values)
    try:
        with db.begin_nested():
            db.execute(stmt)
    except Exception as exc:
        if not _is_psycopg_data_error(exc):
            raise
        cleared = dict(values)
        for k in _SHIPMENT_LINE_DATE_COLS:
            cleared[k] = None
        with db.begin_nested():
            db.execute(_shipment_evidence_line_upsert_statement(cleared))


def _purge_orphan_shipment_evidence_lines(db: Session, job_id: int, seen_source_keys: set[str]) -> int:
    """Remove evidence lines from a prior validate pass whose ``source_key`` no longer appears.

    Runs after a successful re-validate when column mapping changes shift business keys.
    Steward resolution on surviving keys is preserved via upsert-on-conflict semantics.
    """
    if not seen_source_keys:
        return 0
    result = db.execute(
        delete(ShipmentEvidenceLine).where(
            ShipmentEvidenceLine.import_job_id == job_id,
            ShipmentEvidenceLine.source_key.notin_(sorted(seen_source_keys)),
        )
    )
    return int(result.rowcount or 0)


def _flush_shipment_line_batch(db: Session, rows: list[dict[str, Any]]) -> None:
    """Bulk upsert a batch of evidence lines in one statement.

    De-duplicates by ``source_key`` within the batch (keeping the last occurrence — same
    "latest wins" semantics as the previous sequential per-row upsert), so a single
    ``INSERT … ON CONFLICT DO UPDATE`` cannot try to touch the same row twice. On a Postgres
    ``DataError`` (e.g. an out-of-range Excel date anywhere in the batch), falls back to the
    per-row path so the offending row's dates are cleared without losing the rest.
    """
    if not rows:
        return
    deduped: dict[Any, dict[str, Any]] = {}
    for r in rows:
        deduped[r["source_key"]] = r
    batch = list(deduped.values())
    try:
        with db.begin_nested():
            db.execute(_shipment_evidence_line_bulk_upsert_statement(batch))
    except Exception as exc:
        if not _is_psycopg_data_error(exc):
            raise
        for r in batch:
            _execute_shipment_line_upsert(db, r)


def _resolve_unresolved_shipment_lines_for_job(
    db: Session,
    job: ImportJob,
    idx: ProductResolutionIndex,
    source_id: int | None,
    res_cache: "DSIResolutionCache | None" = None,
) -> None:
    """Re-run product and/or distributor resolution only where the corresponding id is still null."""
    ai_enabled = bool(get_settings().ai_assist_enabled)
    jid = int(job.id)
    lines = list(
        db.scalars(
            select(ShipmentEvidenceLine).where(
                ShipmentEvidenceLine.import_job_id == jid,
                or_(ShipmentEvidenceLine.product_id.is_(None), ShipmentEvidenceLine.distributor_id.is_(None)),
            )
        ).all()
    )
    for line in lines:
        if line.product_id is None:
            evidence_date = coalesce_shipment_evidence_date(
                pod_date=line.pod_date,
                est_pod_date=line.est_pod_date,
                promise_date=line.promise_date,
                schedule_ship_date=line.schedule_ship_date,
                ship_confirm_date=line.ship_confirm_date,
                erd_date=line.erd_date,
                exwork_date=line.exwork_date,
            )
            pid, pstatus, ptoken, pdetail = resolve_product_for_evidence(
                idx,
                item_code=line.item_code,
                ean_code=line.ean_code,
                upc_code=line.upc_code,
                sales_model_name=line.sales_model_name,
                evidence_date=evidence_date,
            )
            if ai_enabled and pid is None and pstatus in ("no_match", "ambiguous", "inactive_only", "no_identifier"):
                from app.services.imports.ai_resolver_wiring import (
                    product_candidates_from_index,
                    try_ai_token_resolution,
                )

                raw_prod = line.item_code or line.ean_code or line.upc_code or line.sales_model_name
                ai_id, _ai_tag, ai_suggestion = try_ai_token_resolution(
                    raw_token=raw_prod,
                    token_type="product",
                    candidates=product_candidates_from_index(idx, raw_prod or ""),
                    import_type="shipment_evidence_import",
                    job_id=int(job.id),
                )
                if ai_id is not None:
                    pid = ai_id
                    pstatus = "resolved_unique"
                    pdetail = "ai_auto_resolved"
                elif ai_suggestion is not None:
                    pstatus = "ai_suggested"
                    pdetail = ai_suggestion.reasoning[:256]
            line.product_id = pid
            line.product_resolution_status = pstatus
            line.product_resolution_token = ptoken
            line.product_resolution_detail = pdetail
            db.add(line)
        if line.distributor_id is None:
            did, dstatus, dtoken = resolve_distributor_for_evidence(
                db,
                source_id,
                bill_to=line.bill_to_raw,
                ship_to=line.ship_to_raw,
                res_cache=res_cache,
            )
            if ai_enabled and did is None and dstatus == "unresolved":
                from app.services.imports.ai_resolver_wiring import (
                    distributor_candidates,
                    try_ai_token_resolution,
                )

                dist_tok = line.bill_to_raw or line.ship_to_raw
                ai_id, _ai_tag, ai_suggestion = try_ai_token_resolution(
                    raw_token=dist_tok,
                    token_type="distributor",
                    candidates=distributor_candidates(db, dist_tok or ""),
                    import_type="shipment_evidence_import",
                    job_id=int(job.id),
                )
                if ai_id is not None:
                    did = ai_id
                    dstatus = "resolved"
                    dtoken = dist_tok
                elif ai_suggestion is not None:
                    dstatus = "ai_suggested"
            line.distributor_id = did
            line.distributor_resolution_status = dstatus
            line.distributor_resolution_token = dtoken
            db.add(line)
    db.flush()


def _openpyxl_sheet_to_dataframe(ws: Any) -> pd.DataFrame:
    """Build a DataFrame from a read-only worksheet (values_only, cached numbers when data_only workbook)."""
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return pd.DataFrame()
    if header_row is None:
        return pd.DataFrame()
    header = [str(c).strip() if c is not None else "" for c in header_row]
    data_rows = list(rows_iter)
    if not data_rows:
        return pd.DataFrame(columns=header)
    return pd.DataFrame(list(data_rows), columns=header, dtype=object)


def _load_frames_for_job(job: ImportJob, _df_passed: pd.DataFrame, raw_bytes: bytes) -> list[tuple[str | None, pd.DataFrame, str, str]]:
    """List of (sheet_name, dataframe, report_type, line_state).

    Always reads from ``raw_bytes`` (CSV or XLSX). The pipeline ``df`` argument is ignored so
    multi-sheet XLSX and ``data_only`` values stay consistent with storage.
    """
    fn = job.file_name or ""
    lower = fn.lower()
    out: list[tuple[str | None, pd.DataFrame, str, str]] = []

    if lower.endswith(".csv"):
        df_csv = pd.read_csv(io.BytesIO(raw_bytes))
        cols = _norm_cols(list(df_csv.columns))
        rt, ls = detect_report_type(cols, sheet_name=None, file_name=fn)
        out.append((None, df_csv, rt, ls))
        return out

    if lower.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        try:
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                sdf = _openpyxl_sheet_to_dataframe(ws)
                cols = _norm_cols(list(sdf.columns))
                rt, ls = detect_report_type(cols, sheet_name=sheet, file_name=fn)
                if rt == REPORT_UNKNOWN and sheet.lower() in ("shipped", "unship"):
                    rt = REPORT_ACZA_SHIPPED if sheet.lower() == "shipped" else REPORT_ACZA_UNSHIP
                    ls = LINE_SHIPPED if sheet.lower() == "shipped" else LINE_OPEN_ORDER
                out.append((sheet, sdf, rt, ls))
        finally:
            wb.close()
        return out

    df_fallback = _df_passed if isinstance(_df_passed, pd.DataFrame) and not _df_passed.empty else pd.DataFrame()
    cols = _norm_cols(list(df_fallback.columns))
    rt, ls = detect_report_type(cols, sheet_name=None, file_name=fn)
    out.append((None, df_fallback, rt, ls))
    return out


def _src_to_canonical_rev(src_to_canon: dict[str, Any] | None) -> dict[str, str] | None:
    """Invert file-header → canonical mapping to canonical → file header (first header wins per canonical)."""
    if not src_to_canon:
        return None
    rev: dict[str, str] = {}
    for src, can in src_to_canon.items():
        if not isinstance(src, str) or not isinstance(can, str):
            continue
        s = src.strip()
        c = can.strip()
        if not s or not c:
            continue
        if c not in rev:
            rev[c] = s
    return rev or None


def process_shipment_evidence_import(
    db: Session,
    job: ImportJob,
    df: pd.DataFrame,
    mapping: dict[str, str],
    on_progress: Any = None,
) -> int:
    """Parse file(s), write ``ShipmentEvidenceLine`` rows. Returns blocking error count."""
    effective_src_to_canon: dict[str, Any] = dict(job.field_mapping or mapping or {})
    header_by_canonical = _src_to_canonical_rev(effective_src_to_canon)
    raw_meta = db.scalars(select(RawFileMetadata).where(RawFileMetadata.job_id == job.id)).first()
    if not raw_meta:
        db.add(
            ImportRowResult(
                job_id=job.id,
                row_number=0,
                severity="error",
                code="missing_raw_file",
                message="No raw file metadata for this import job.",
            )
        )
        return 1

    storage = get_storage_backend()
    raw_bytes = storage.read(raw_meta.storage_key)
    frames = _load_frames_for_job(job, df, raw_bytes)

    idx = _load_product_resolution_index(db)
    source_id = int(job.source_id) if job.source_id else None
    # Pre-load distributor master + approved aliases once so per-row resolution is in-memory
    # (replaces a full dim_distributor table scan per row — the dominant validate cost).
    res_cache = _build_distributor_resolution_cache(db, source_id)
    ai_enabled = bool(get_settings().ai_assist_enabled)

    total_rows = 0
    for _sn, _frame, _rt, _ls in frames:
        if _frame is not None and _rt != REPORT_UNKNOWN:
            total_rows += len(_frame)

    line_buffer: list[dict[str, Any]] = []
    seen_source_keys: set[str] = set()

    def _flush_buffer() -> None:
        if line_buffer:
            for row in line_buffer:
                sk = row.get("source_key")
                if isinstance(sk, str) and sk:
                    seen_source_keys.add(sk)
            _flush_shipment_line_batch(db, line_buffer)
            line_buffer.clear()

    source = job.source
    if source and isinstance(source.column_mapping_memory, dict) and frames:
        from app.services.imports.ai_resolver_wiring import record_format_drift_on_job

        first_headers: list[str] = []
        for _sn, frame, _rt, _ls in frames:
            if frame is not None and len(frame.columns):
                first_headers = [str(c) for c in frame.columns]
                break
        if first_headers:
            record_format_drift_on_job(
                job,
                current_headers=first_headers,
                column_mapping_memory=source.column_mapping_memory,
                field_mapping=effective_src_to_canon,
            )

    blocking = 0
    global_row = 0
    unknown_reports = 0

    for sheet_name, frame, report_type, line_state in frames:
        if frame is None or len(frame) == 0:
            continue
        if report_type == REPORT_UNKNOWN:
            unknown_reports += 1
            continue

        for pos, (_, row) in enumerate(frame.iterrows(), start=2):
            global_row += 1
            try:
                series = row if isinstance(row, pd.Series) else pd.Series(row, index=frame.columns)
                raw_payload = _row_dict(series)
                ex = _extract_common(series, header_by_canonical)
                source_key = stable_source_key_for_row(report_type=report_type, sheet_name=sheet_name, ex=ex)

                evidence_date = coalesce_shipment_evidence_date(
                    pod_date=ex["pod_date"],
                    est_pod_date=ex["est_pod_date"],
                    promise_date=ex["promise_date"],
                    schedule_ship_date=ex["schedule_ship_date"],
                    ship_confirm_date=ex["ship_confirm_date"],
                    erd_date=ex["erd_date"],
                    exwork_date=ex["exwork_date"],
                )
                pid, pstatus, ptoken, pdetail = resolve_product_for_evidence(
                    idx,
                    item_code=ex["item_code"],
                    ean_code=ex["ean_code"],
                    upc_code=ex["upc_code"],
                    sales_model_name=ex["sales_model_name"],
                    evidence_date=evidence_date,
                )
                did, dstatus, dtoken = resolve_distributor_for_evidence(
                    db,
                    source_id,
                    bill_to=ex["bill_to_raw"],
                    ship_to=ex["ship_to_raw"],
                    res_cache=res_cache,
                )

                if ai_enabled and pid is None and pstatus in ("no_match", "ambiguous", "inactive_only", "no_identifier"):
                    from app.services.imports.ai_resolver_wiring import (
                        product_candidates_from_index,
                        stash_ai_suggestion_on_payload,
                        try_ai_token_resolution,
                    )

                    raw_prod = ex["item_code"] or ex["ean_code"] or ex["upc_code"] or ex["sales_model_name"]
                    ai_id, ai_tag, ai_suggestion = try_ai_token_resolution(
                        raw_token=raw_prod,
                        token_type="product",
                        candidates=product_candidates_from_index(idx, raw_prod or ""),
                        import_type="shipment_evidence_import",
                        job_id=int(job.id),
                    )
                    if ai_id is not None:
                        pid = ai_id
                        pstatus = "resolved_unique"
                        pdetail = "ai_auto_resolved"
                    elif ai_suggestion is not None:
                        pstatus = "ai_suggested"
                        pdetail = ai_suggestion.reasoning[:256]
                        raw_payload = stash_ai_suggestion_on_payload(
                            raw_payload, token_type="product", suggestion=ai_suggestion
                        )

                if ai_enabled and did is None and dstatus == "unresolved":
                    from app.services.imports.ai_resolver_wiring import (
                        distributor_candidates,
                        stash_ai_suggestion_on_payload,
                        try_ai_token_resolution,
                    )

                    dist_tok = ex["bill_to_raw"] or ex["ship_to_raw"]
                    ai_id, ai_tag, ai_suggestion = try_ai_token_resolution(
                        raw_token=dist_tok,
                        token_type="distributor",
                        candidates=distributor_candidates(db, dist_tok or ""),
                        import_type="shipment_evidence_import",
                        job_id=int(job.id),
                    )
                    if ai_id is not None:
                        did = ai_id
                        dstatus = "resolved"
                        dtoken = dist_tok
                    elif ai_suggestion is not None:
                        dstatus = "ai_suggested"
                        raw_payload = stash_ai_suggestion_on_payload(
                            raw_payload, token_type="distributor", suggestion=ai_suggestion
                        )

                q_dec = _decimal_or_none(ex["quantity"])
                row_values: dict[str, Any] = {
                    "import_job_id": int(job.id),
                    "source_key": source_key,
                    "source_sheet": sheet_name,
                    "source_row_number": pos,
                    "report_type": report_type,
                    "line_state": line_state,
                    "raw_source_row": raw_payload,
                    "operating_unit": ex["operating_unit"],
                    "bill_to_raw": ex["bill_to_raw"],
                    "ship_to_raw": ex["ship_to_raw"],
                    "order_no": ex["order_no"],
                    "customer_po": ex["customer_po"],
                    "order_line": ex["order_line"],
                    "delivery_no": ex["delivery_no"],
                    "invoice_line": ex["invoice_line"],
                    "item_code": ex["item_code"],
                    "sales_model_name": ex["sales_model_name"],
                    "customer_item": ex["customer_item"],
                    "ean_code": ex["ean_code"],
                    "upc_code": ex["upc_code"],
                    "mpor_item_no": ex["mpor_item_no"],
                    "quantity": float(q_dec) if q_dec is not None else None,
                    "unit_price": float(v) if (v := _decimal_or_none(ex["unit_price"])) is not None else None,
                    "amount": float(v) if (v := _decimal_or_none(ex["amount"])) is not None else None,
                    "currency_code": ex["currency_code"],
                    "ship_confirm_date": ex["ship_confirm_date"],
                    "schedule_ship_date": ex["schedule_ship_date"],
                    "promise_date": ex["promise_date"],
                    "exwork_date": ex["exwork_date"],
                    "erd_date": ex["erd_date"],
                    "est_pod_date": ex["est_pod_date"],
                    "pod_date": ex["pod_date"],
                    "customer_dealer_token": ex["customer_dealer_token"],
                    "customer_resolution_status": None,
                    "customer_id": None,
                    "product_id": pid,
                    "product_resolution_status": pstatus,
                    "product_resolution_token": ptoken,
                    "product_resolution_detail": pdetail,
                    "distributor_id": did,
                    "distributor_resolution_status": dstatus,
                    "distributor_resolution_token": dtoken,
                }
                line_buffer.append(row_values)
            except ShipmentEvidenceSourceKeyError as exc:
                blocking += 1
                db.add(
                    ImportRowResult(
                        job_id=job.id,
                        row_number=global_row,
                        severity="error",
                        code="shipment_evidence_source_key",
                        message=str(exc)[:2000],
                        raw_payload={"sheet": sheet_name, "row_index": pos, "report_type": report_type},
                    )
                )
            except Exception as exc:  # noqa: BLE001
                blocking += 1
                db.add(
                    ImportRowResult(
                        job_id=job.id,
                        row_number=global_row,
                        severity="error",
                        code="shipment_evidence_row_error",
                        message=str(exc)[:2000],
                        raw_payload={"sheet": sheet_name, "row_index": pos},
                    )
                )

            if len(line_buffer) >= _SHIPMENT_UPSERT_CHUNK:
                _flush_buffer()
                if on_progress is not None:
                    on_progress("writing_shipment_lines", "Writing shipment evidence", global_row, total_rows)

    _flush_buffer()
    if on_progress is not None:
        on_progress("writing_shipment_lines", "Writing shipment evidence", global_row, total_rows)

    db.flush()
    _resolve_unresolved_shipment_lines_for_job(db, job, idx, source_id, res_cache=res_cache)

    if unknown_reports == len(frames) and global_row == 0:
        blocking += 1
        db.add(
            ImportRowResult(
                job_id=job.id,
                row_number=0,
                severity="error",
                code="shipment_evidence_unknown_format",
                message="Could not detect a supported shipment / order report from headers.",
            )
        )

    if blocking == 0 and seen_source_keys:
        _purge_orphan_shipment_evidence_lines(db, int(job.id), seen_source_keys)

    db.flush()
    from app.services.imports.shipment_purchase_order_materialize import materialize_purchase_orders_for_shipment_job

    materialize_purchase_orders_for_shipment_job(db, int(job.id))
    from app.services.imports.shipment_evidence_observations import sync_job_observations_after_validate

    sync_job_observations_after_validate(db, job)
    _rebuild_shipment_distributor_candidates(db, job)
    _rebuild_shipment_customer_candidates(db, job)

    meta = dict(job.staged_metadata or {})
    if not str(meta.get("import_purpose") or "").strip():
        meta["import_purpose"] = "current"
    meta["shipment_evidence"] = to_jsonable(
        {
            "processed_at": datetime.now(timezone.utc).isoformat(),
            "sheets": [s for s, *_ in frames],
            "total_lines_written_estimate": global_row,
        }
    )
    job.staged_metadata = to_jsonable(meta)

    summary = {
        "lines": global_row,
        "blocking": blocking,
        "sheets": len(frames),
    }
    db.add(
        ImportRowResult(
            job_id=job.id,
            row_number=0,
            severity="info" if blocking == 0 else "warning",
            code="shipment_evidence_summary",
            message=json.dumps(summary),
        )
    )
    return 1 if blocking else 0
