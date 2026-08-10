"""B2 — export + apply helpers for lineup net-requirement."""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.commercial_lineup import CommercialLineupCase, CommercialLineupLine
from app.models.dimensions import DimCustomer, DimProduct
from app.services.lineup.bias_correction import volume_bias_by_bu
from app.services.lineup.bulk import bulk_upsert_lineup_items
from app.services.lineup.net_requirement import build_net_requirement_rows

NET_REQUIREMENT_CSV_HEADERS = [
    "distributor_id",
    "product_id",
    "business_unit",
    "period_from",
    "period_to",
    "forecast_demand",
    "bias_adjusted_forecast",
    "channel_stock",
    "in_transit",
    "target_cover_units",
    "net_requirement",
    "weekly_velocity",
    "bias_factor",
]

# Tenant workbook on-ramp sheet headers (generic labels — tenant may remap via profile sheet names)
TENANT_LINEUP_SHEET_HEADERS = [
    "Customer Code",
    "Customer Name",
    "SKU",
    "Product Name",
    "Period Label",
    "Period Start",
    "Planned Qty",
    "Distributor ID",
    "Product ID",
    "Business Unit",
    "Forecast Demand",
    "Bias Adjusted Forecast",
    "Channel Stock",
    "In Transit",
    "Target Cover",
    "Net Requirement",
    "Notes",
]


def net_requirement_to_csv(payload: dict[str, Any]) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(NET_REQUIREMENT_CSV_HEADERS)
    for r in payload.get("rows") or []:
        writer.writerow(
            [
                r.get("distributor_id"),
                r.get("product_id"),
                r.get("business_unit") or "",
                r.get("period_from"),
                r.get("period_to"),
                r.get("forecast_demand"),
                r.get("bias_adjusted_forecast"),
                r.get("channel_stock"),
                r.get("in_transit"),
                r.get("target_cover_units"),
                r.get("net_requirement"),
                r.get("weekly_velocity"),
                r.get("bias_factor"),
            ]
        )
    return buf.getvalue()


def net_requirement_to_xlsx_bytes(
    payload: dict[str, Any],
    *,
    draft_rows: list[dict[str, Any]] | None = None,
    net_requirement_sheet: str | None = None,
    draft_lineup_sheet: str | None = None,
) -> bytes:
    """Tenant workbook on-ramp: NetRequirement + optional DraftLineup sheets.

    Sheet titles come from tenant profile (defaults: NetRequirement / DraftLineup) —
    never OEM-branded application law.
    """
    from app.services.commercial_tenant_profile import lineup_export_sheet_names

    sheets = lineup_export_sheet_names()
    nr_title = (net_requirement_sheet or sheets["net_requirement"])[:31] or "NetRequirement"
    draft_title = (draft_lineup_sheet or sheets["draft_lineup"])[:31] or "DraftLineup"

    wb = Workbook()
    ws = wb.active
    ws.title = nr_title
    ws.append(list(NET_REQUIREMENT_CSV_HEADERS))
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in payload.get("rows") or []:
        ws.append(
            [
                r.get("distributor_id"),
                r.get("product_id"),
                r.get("business_unit") or "",
                r.get("period_from"),
                r.get("period_to"),
                r.get("forecast_demand"),
                r.get("bias_adjusted_forecast"),
                r.get("channel_stock"),
                r.get("in_transit"),
                r.get("target_cover_units"),
                r.get("net_requirement"),
                r.get("weekly_velocity"),
                r.get("bias_factor"),
            ]
        )

    draft = list(draft_rows or [])
    if draft:
        ws2 = wb.create_sheet(draft_title)
        ws2.append(list(TENANT_LINEUP_SHEET_HEADERS))
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        ws2.freeze_panes = "A2"
        for row in draft:
            ws2.append([row.get(h) for h in TENANT_LINEUP_SHEET_HEADERS])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def half_year_period_starts(year: int, half: int) -> list[tuple[date, str]]:
    if half not in (1, 2):
        raise ValueError("half must be 1 or 2")
    if half == 1:
        return [(date(year, 1, 1), f"{str(year)[2:]}Q1"), (date(year, 4, 1), f"{str(year)[2:]}Q2")]
    return [(date(year, 7, 1), f"{str(year)[2:]}Q3"), (date(year, 10, 1), f"{str(year)[2:]}Q4")]


def _normalize_period_label_display(period_label: str | None, period_start: date) -> str:
    if period_label and str(period_label).strip():
        return str(period_label).strip()
    q = (period_start.month - 1) // 3 + 1
    return f"{str(period_start.year)[2:]}Q{q}"


async def _product_bu_map(db: AsyncSession) -> dict[int, str]:
    product_bu: dict[int, str] = {}
    for pid, pl, bu in (
        await db.execute(select(DimProduct.id, DimProduct.product_line, DimProduct.business_unit))
    ).all():
        label = (bu or pl or "").strip()
        if label:
            product_bu[int(pid)] = label
    return product_bu


async def _bias_by_bu(db: AsyncSession, *, apply_bias: bool) -> dict[str, float]:
    if not apply_bias:
        return {}
    bias_meta = await volume_bias_by_bu(db)
    return dict(bias_meta.get("by_bu") or {})


async def apply_net_requirement_to_lineup(
    db: AsyncSession,
    *,
    period_start: date,
    period_label: str | None = None,
    distributor_id: int | None = None,
    horizon_weeks: int = 13,
    target_cover_weeks: float = 4.0,
    replace_matching: bool = True,
    limit: int = 200,
    apply_bias: bool = False,
    write_commercial_case: bool = True,
) -> dict[str, Any]:
    product_bu = await _product_bu_map(db)
    bias_by_bu = await _bias_by_bu(db, apply_bias=apply_bias)

    payload = await build_net_requirement_rows(
        db,
        horizon_weeks=horizon_weeks,
        target_cover_weeks=target_cover_weeks,
        distributor_id=distributor_id,
        product_bu=product_bu,
        bias_by_bu=bias_by_bu,
        include_customer_shares=True,
        limit=limit,
    )

    cust_by_id = {
        int(r.id): r for r in (await db.execute(select(DimCustomer))).scalars().all()
    }
    prod_by_id = {
        int(r.id): r for r in (await db.execute(select(DimProduct))).scalars().all()
    }

    label = _normalize_period_label_display(period_label, period_start)
    rows: list[dict[str, Any]] = []
    commercial_lines: list[dict[str, Any]] = []
    skipped_zero = 0
    skipped_no_alloc = 0
    for nr in payload.get("rows") or []:
        net = float(nr.get("net_requirement") or 0)
        if net <= 0:
            skipped_zero += 1
            continue
        alloc = nr.get("customer_allocation") or []
        if not alloc:
            skipped_no_alloc += 1
            continue
        pid = int(nr["product_id"])
        prod = prod_by_id.get(pid)
        if not prod or not prod.sku:
            continue
        for a in alloc:
            cid = int(a["customer_id"])
            cust = cust_by_id.get(cid)
            if not cust or not cust.code:
                continue
            units = float(a.get("suggested_lineup_units") or 0)
            if units <= 0:
                continue
            note = f"b2_net_requirement dist={nr['distributor_id']} forecast_share={a.get('share')}"
            rows.append(
                {
                    "customer_code": str(cust.code),
                    "channel_code": None,
                    "period_start": period_start.isoformat(),
                    "period_label": label,
                    "sku": str(prod.sku),
                    "planned_volume_units": units,
                    "approval_status": "draft",
                    "notes": note,
                }
            )
            commercial_lines.append(
                {
                    "product_id": pid,
                    "customer_id": cid,
                    "distributor_id": int(nr["distributor_id"]),
                    "quantity_units": units,
                    "sku_raw": str(prod.sku),
                    "customer_token": str(cust.code),
                    "notes": note,
                    "business_unit": nr.get("business_unit"),
                    "forecast_demand": nr.get("forecast_demand"),
                    "bias_adjusted_forecast": nr.get("bias_adjusted_forecast"),
                    "channel_stock": nr.get("channel_stock"),
                    "in_transit": nr.get("in_transit"),
                    "target_cover_units": nr.get("target_cover_units"),
                    "net_requirement": nr.get("net_requirement"),
                    "customer_name": cust.name,
                    "product_name": prod.name,
                }
            )

    if not rows:
        return {
            "inserted": 0,
            "updated": 0,
            "skipped": 0,
            "errors": 0,
            "results": [],
            "skipped_zero_net": skipped_zero,
            "skipped_no_allocation": skipped_no_alloc,
            "source_row_count": payload.get("row_count", 0),
            "draft_rows_built": 0,
            "apply_bias": bool(apply_bias and bias_by_bu),
            "commercial_case_id": None,
            "commercial_lines_written": 0,
        }

    out = await bulk_upsert_lineup_items(db, rows, replace_matching=replace_matching)
    out["skipped_zero_net"] = skipped_zero
    out["skipped_no_allocation"] = skipped_no_alloc
    out["source_row_count"] = payload.get("row_count", 0)
    out["draft_rows_built"] = len(rows)
    out["apply_bias"] = bool(apply_bias and bias_by_bu)

    case_id: int | None = None
    lines_written = 0
    if write_commercial_case and commercial_lines:
        case = CommercialLineupCase(
            period_label=label,
            inferred_period_start=period_start,
            commercial_status="draft_imported",
            import_intent="current_working_lineup",
            source_context="lineup_net_requirement",
            file_name="b2_net_requirement_apply.xlsx",
            notes=f"B2 apply net-requirement period_start={period_start.isoformat()} bias={bool(apply_bias and bias_by_bu)}",
        )
        db.add(case)
        await db.flush()
        case_id = int(case.id)
        for cl in commercial_lines:
            db.add(
                CommercialLineupLine(
                    case_id=case_id,
                    product_id=cl["product_id"],
                    customer_id=cl["customer_id"],
                    distributor_id=cl["distributor_id"],
                    quantity_units=cl["quantity_units"],
                    sku_raw=cl["sku_raw"],
                    customer_token=cl["customer_token"],
                    internal_notes=cl["notes"],
                    row_status="imported",
                    diagnostic_codes=["b2_net_requirement_apply"],
                )
            )
            lines_written += 1
        await db.commit()

    out["commercial_case_id"] = case_id
    out["commercial_lines_written"] = lines_written
    return out


async def build_tenant_workbook_export(
    db: AsyncSession,
    *,
    period_start: date | None = None,
    period_label: str | None = None,
    distributor_id: int | None = None,
    horizon_weeks: int = 13,
    target_cover_weeks: float = 4.0,
    apply_bias: bool = False,
    limit: int = 500,
) -> bytes:
    """XLSX on-ramp: NetRequirement sheet + DraftLineup projection from net req allocation."""
    product_bu = await _product_bu_map(db)
    bias_by_bu = await _bias_by_bu(db, apply_bias=apply_bias)
    payload = await build_net_requirement_rows(
        db,
        horizon_weeks=horizon_weeks,
        target_cover_weeks=target_cover_weeks,
        distributor_id=distributor_id,
        product_bu=product_bu,
        bias_by_bu=bias_by_bu,
        include_customer_shares=True,
        limit=limit,
    )

    cust_by_id = {
        int(r.id): r for r in (await db.execute(select(DimCustomer))).scalars().all()
    }
    prod_by_id = {
        int(r.id): r for r in (await db.execute(select(DimProduct))).scalars().all()
    }
    start = period_start or date.today().replace(day=1)
    label = _normalize_period_label_display(period_label, start)

    draft_rows: list[dict[str, Any]] = []
    for nr in payload.get("rows") or []:
        net = float(nr.get("net_requirement") or 0)
        if net <= 0:
            continue
        pid = int(nr["product_id"])
        prod = prod_by_id.get(pid)
        for a in nr.get("customer_allocation") or []:
            units = float(a.get("suggested_lineup_units") or 0)
            if units <= 0:
                continue
            cust = cust_by_id.get(int(a["customer_id"]))
            draft_rows.append(
                {
                    "Customer Code": cust.code if cust else "",
                    "Customer Name": cust.name if cust else "",
                    "SKU": prod.sku if prod else "",
                    "Product Name": prod.name if prod else "",
                    "Period Label": label,
                    "Period Start": start.isoformat(),
                    "Planned Qty": units,
                    "Distributor ID": nr.get("distributor_id"),
                    "Product ID": pid,
                    "Business Unit": nr.get("business_unit") or "",
                    "Forecast Demand": nr.get("forecast_demand"),
                    "Bias Adjusted Forecast": nr.get("bias_adjusted_forecast"),
                    "Channel Stock": nr.get("channel_stock"),
                    "In Transit": nr.get("in_transit"),
                    "Target Cover": nr.get("target_cover_units"),
                    "Net Requirement": nr.get("net_requirement"),
                    "Notes": "b2_tenant_workbook_onramp",
                }
            )

    return net_requirement_to_xlsx_bytes(payload, draft_rows=draft_rows)
